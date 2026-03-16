"""doctool 固有テストアダプタ

Excel設計書レビュー指摘事項抽出ツール（doctool）に特化したテストアダプタ。
汎用テストエンジンの ToolAdapter Protocol を実装し、doctool 固有の以下の処理を提供する:

- メインマクロエントリポイントの特定（Sheet1.CmdGen_Click_Core）
- xlsm 事前設定（基本設定シート / 指摘分類マッピング設定シート 操作）
- テンプレートシートのアサーション評価
- 削除系マクロ実行（コメント削除 / シート削除）

このファイルは vba-text-based-dev/test-framework/scripts/adapters/ 配下に配置する。
doctool/test/ 配下の tool_config.yaml から動的ロードされる。
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

import openpyxl

from adapters import BaseToolAdapter, ComparisonConfig

# ダイアログログファイル名（scenario_runner.py と共有）
_DIALOG_LOG_FILENAME = "dialog_log.txt"


class DoctoolAdapter(BaseToolAdapter):
    """Excel設計書レビュー指摘事項抽出ツール用テストアダプタ。

    doctool 固有の VBA マクロ・シート構造・設定処理を汎用テストエンジンから分離する。
    """

    XLSM_NAME = "Excel設計書レビュー指摘事項抽出ツール.xlsm"

    # =====================================================================
    # ToolAdapter Protocol 実装
    # =====================================================================

    def get_macro_entry_point(self) -> str:
        """doctool のメインマクロパスを返す。"""
        return "Sheet1.CmdGen_Click_Core"

    def get_default_comparison(self) -> ComparisonConfig:
        """doctool のデフォルト比較設定を返す。

        印刷範囲（print_area）は doctool では比較対象外とする。
        """
        return ComparisonConfig(print_area=False)

    def get_delete_macro_path(self, action: str) -> str:
        """削除系アクションに対応する VBA マクロパスを返す。

        Args:
            action: "delete_comments" または "delete_sheets"

        Returns:
            VBA マクロのフルパス
        """
        paths = {
            "delete_comments": "Module2.DelAllReviewComments_Click_Core",
            "delete_sheets": "Module2.DelAllReviewResultSheets_Click_Core",
        }
        return paths[action]

    def get_clear_log_macro(self) -> str:
        """ダイアログログをクリアする VBA マクロパスを返す。"""
        return "Module2.ClearDialogLog"

    def get_dialog_log_vba_function(self) -> str:
        """ダイアログログを取得する VBA 関数パスを返す。"""
        return "Module2.GetDialogLog"

    def apply_setup(
        self,
        scenario_name: str,
        xlsm_wb: Any,
        setup: Dict[str, Any],
        work_dir: Path,
    ) -> None:
        """xlsm の事前設定を適用する（doctool 固有）。

        config.yaml の setup セクションに対応する以下の設定を行う:
        - use_review_record: 基本設定シート B2
        - use_summary: 基本設定シート B3
        - review_list_file: REVIEW_LIST_FILEPATH 名前付き範囲
        - categories: 指摘分類マッピング設定シート
        - named_ranges: 任意の名前付き範囲
        - item_mapping_cells: 項目マッピング設定シートの個別セル
        """
        settings_ws = xlsm_wb.sheets["基本設定"]

        if "use_review_record" in setup:
            settings_ws["B2"].value = setup["use_review_record"]
            print(f"[{scenario_name}] setup: use_review_record={setup['use_review_record']}")

        if "use_summary" in setup:
            settings_ws["B3"].value = setup["use_summary"]
            print(f"[{scenario_name}] setup: use_summary={setup['use_summary']}")

        if "review_list_file" in setup:
            review_list_path = work_dir / setup["review_list_file"]
            xlsm_wb.names["REVIEW_LIST_FILEPATH"].refers_to_range.value = str(review_list_path)
            print(f"[{scenario_name}] setup: REVIEW_LIST_FILEPATH={review_list_path}")

        if "categories" in setup:
            self._apply_categories(scenario_name, xlsm_wb, setup["categories"])

        if "named_ranges" in setup:
            for range_name, value in setup["named_ranges"].items():
                try:
                    xlsm_wb.names[range_name].refers_to_range.value = value
                    print(f"[{scenario_name}] setup: named_ranges[{range_name}]={repr(value)}")
                except Exception as e:
                    print(f"[{scenario_name}] setup: named_ranges warning [{range_name}]: {e}")

        if "controls" in setup:
            # VBA コード名 "Sheet1" のシートをタブ名によらず検索する
            # xlsm_wb.sheets["Sheet1"] はタブ名で検索するため、タブ名が異なると
            # pywintypes.com_error (-2147352567) が発生する
            target_sheet = None
            for sheet in xlsm_wb.sheets:
                try:
                    if sheet.api.CodeName == "Sheet1":
                        target_sheet = sheet
                        break
                except Exception:
                    pass
            if target_sheet is None:
                print(f"[{scenario_name}] setup: controls warning: VBA コード名 'Sheet1' のシートが見つかりません")
            else:
                for control_name, value in setup["controls"].items():
                    try:
                        target_sheet.api.OLEObjects(control_name).Object.Value = value
                        print(f"[{scenario_name}] setup: controls[{control_name}]={repr(value)}")
                    except Exception as e:
                        print(f"[{scenario_name}] setup: controls warning [{control_name}]: {e}")

        if "item_mapping_cells" in setup:
            map_ws = xlsm_wb.sheets["項目マッピング設定"]
            for cell_ref, value in setup["item_mapping_cells"].items():
                map_ws[cell_ref].value = value
                print(f"[{scenario_name}] setup: item_mapping_cells[{cell_ref}]={repr(value)}")

    def pre_macro_hook(
        self,
        scenario_name: str,
        xlsm_wb: Any,
        step: Dict[str, Any],
        work_dir: Path,
    ) -> None:
        """extract ステップのマクロ実行前処理（ステップ固有カテゴリ設定）。"""
        if step.get("action") == "extract" and "categories" in step:
            self._apply_categories(scenario_name, xlsm_wb, step["categories"])

    def post_macro_hook(
        self,
        scenario_name: str,
        xlsm_wb: Any,
        step: Dict[str, Any],
        work_dir: Path,
    ) -> None:
        """マクロ実行後処理（現時点では何もしない）。"""
        pass

    def evaluate_custom_assertions(
        self,
        work_dir: Path,
        assertions: List[Dict[str, Any]],
    ) -> List[str]:
        """テンプレートシートのアサーションを評価する（doctool 固有）。

        config.yaml の template_assertions セクションに対応する以下の検証を行う:
        - category_count: 行7 B列以降の非空セル数
        - cell_formula_contains: 指定セルの数式が特定文字列を含む
        """
        return _evaluate_template_assertions(work_dir, assertions, self.XLSM_NAME)

    def execute_steps(
        self,
        scenario_name: str,
        xlsm_wb: Any,
        steps: List[Dict[str, Any]],
        test_mode: bool,
        work_dir: Path,
        platform: Any,
    ) -> None:
        """doctool の全ステップを実行する。

        extract / delete_comments / delete_sheets の各アクションを処理する。
        scenario_runner.py の _execute_vba() のアクションループと同等の処理を行う。

        Args:
            scenario_name: ログ用シナリオ名
            xlsm_wb: xlwings Workbook オブジェクト（xlsm）
            steps: config.yaml の steps リスト
            test_mode: testMode フラグ（True=自動、False=手動）
            work_dir: テスト作業ディレクトリ
            platform: ExcelPlatform インスタンス（マクロリトライ実行に使用）
        """
        from helpers.config_loader import validate_step

        macro = xlsm_wb.macro(self.get_macro_entry_point())

        for step_idx, step in enumerate(steps, 1):
            validate_step(step)
            action = step["action"]

            if action == "extract":
                # マクロ実行前処理（ステップ固有カテゴリ設定等）
                self.pre_macro_hook(scenario_name, xlsm_wb, step, work_dir)

                review_times = step["review_times"]
                repeat = step.get("repeat", 1)
                print(
                    f"[{scenario_name}] Step {step_idx}: extract"
                    f" (REVIEW_TIMES={review_times}, repeat={repeat})"
                )
                for run_num in range(repeat):
                    xlsm_wb.names["REVIEW_TIMES"].refers_to_range.value = review_times
                    if test_mode:
                        xlsm_wb.macro(self.get_clear_log_macro())()
                    print(f"[{scenario_name}]   マクロ実行中 (run {run_num + 1}/{repeat})...")
                    platform.run_macro_with_retry(scenario_name, macro, test_mode)
                    if test_mode:
                        self._flush_dialog_log(scenario_name, xlsm_wb, work_dir)

                # マクロ実行後処理
                self.post_macro_hook(scenario_name, xlsm_wb, step, work_dir)

            elif action in ("delete_comments", "delete_sheets"):
                print(f"[{scenario_name}] Step {step_idx}: {action}")
                try:
                    xlsm_wb.macro(self.get_clear_log_macro())()
                    xlsm_wb.macro(self.get_delete_macro_path(action))(test_mode)
                    self._flush_dialog_log(scenario_name, xlsm_wb, work_dir)
                except Exception as e:
                    print(f"[{scenario_name}]   Warning: {e}")

    def teardown(self) -> None:
        pass

    # =====================================================================
    # doctool 固有の内部メソッド
    # =====================================================================

    def _flush_dialog_log(self, scenario_name: str, xlsm_wb: Any, work_dir: Path) -> None:
        """VBA ダイアログログを dialog_log.txt に追記し、コンソールにも出力する。"""
        dialog_log = xlsm_wb.macro(self.get_dialog_log_vba_function())()
        if dialog_log:
            log_path = work_dir / _DIALOG_LOG_FILENAME
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(dialog_log + "\n")
            for line in dialog_log.split("\n"):
                if line:
                    print(f"[{scenario_name}]   [LOG] {line}")

    def _run_delete_macro(
        self,
        scenario_name: str,
        xlsm_wb: Any,
        macro_path: str,
        test_mode: bool,
        work_dir: Path,
    ) -> None:
        """削除系マクロを実行する（doctool 固有）。

        コメント削除（Module2.DelAllReviewComments_Click_Core）と
        シート削除（Module2.DelAllReviewResultSheets_Click_Core）で使用する。

        Args:
            scenario_name: ログ用シナリオ名
            xlsm_wb: xlwings Workbook オブジェクト
            macro_path: 実行するマクロのフルパス
            test_mode: testMode フラグ（True=自動、False=手動）
            work_dir: テスト作業ディレクトリ
        """
        try:
            xlsm_wb.macro("Module2.ClearDialogLog")()
            xlsm_wb.macro(macro_path)(test_mode)
            self._flush_dialog_log(scenario_name, xlsm_wb, work_dir)
        except Exception as e:
            print(f"[{scenario_name}]   Warning: {e}")

    def _apply_categories(
        self,
        scenario_name: str,
        xlsm_wb: Any,
        categories: List[Dict[str, Any]],
    ) -> None:
        """指摘分類マッピング設定シートをカテゴリリストで上書きする。

        Args:
            scenario_name: ログ用シナリオ名
            xlsm_wb: xlwings Workbook オブジェクト
            categories: [{"alias": "a", "name": "01_要件漏れ"}, ...] のリスト
        """
        cat_ws = xlsm_wb.sheets["指摘分類マッピング設定"]
        last_row = cat_ws.range("A1").current_region.last_cell.row
        if last_row >= 2:
            cat_ws.range(f"A2:B{last_row}").clear_contents()
        for i, cat in enumerate(categories, start=2):
            cat_ws.range(f"A{i}").value = cat["alias"]
            cat_ws.range(f"B{i}").value = cat["name"]
        aliases = [c["alias"] for c in categories]
        print(f"[{scenario_name}] categories: {len(categories)} カテゴリを設定 ({aliases})")


# ---------------------------------------------------------------------------
# モジュールレベルヘルパー（openpyxl 依存、xlwings 不要で WSL2 から実行可能）
# ---------------------------------------------------------------------------

def _evaluate_template_assertions(
    work_dir: Path,
    assertions: List[Dict[str, Any]],
    xlsm_name: str,
) -> List[str]:
    """テンプレートシートのアサーションを評価し、エラーメッセージのリストを返す。

    VBA 実行・保存後の xlsm ファイルを openpyxl で読み込み、以下を検証する:
    - category_count: 行7 B列以降の非空セル数（カテゴリ列数）
    - cell_formula_contains: 指定セルの数式が特定文字列を含む

    Args:
        work_dir: VBA 実行後の xlsm が格納されたディレクトリ
        assertions: config.yaml の template_assertions リスト
        xlsm_name: xlsm ファイル名

    Returns:
        エラーメッセージのリスト（空なら全アサーション通過）
    """
    xlsm_path = work_dir / xlsm_name
    errors: List[str] = []

    try:
        wb_values = openpyxl.load_workbook(str(xlsm_path), keep_vba=True, data_only=True)
        wb_formulas = openpyxl.load_workbook(str(xlsm_path), keep_vba=True, data_only=False)
    except Exception as e:
        return [f"template_assertions: xlsm の読み込みに失敗しました: {e}"]

    for assertion in assertions:
        sheet_name = assertion.get("sheet", "レビュー結果シートテンプレート")

        if sheet_name not in wb_values.sheetnames:
            errors.append(f"template_assertions: シート '{sheet_name}' が存在しません")
            continue

        ws_v = wb_values[sheet_name]
        ws_f = wb_formulas[sheet_name]

        # カテゴリ列数チェック（行7 B列以降の非空セル数）
        if "category_count" in assertion:
            expected = assertion["category_count"]
            count = 0
            col = 2  # B列
            while ws_v.cell(7, col).value is not None:
                count += 1
                col += 1
            if count != expected:
                errors.append(
                    f"[{sheet_name}] category_count: expected={expected}, actual={count}"
                )
            else:
                print(f"  ✓ [{sheet_name}] category_count = {count}")

        # 数式の文字列チェック
        for fc in assertion.get("cell_formula_contains", []):
            cell_ref = fc["cell"]
            expected_text = fc["contains"]
            formula = ws_f[cell_ref].value
            if formula is None:
                errors.append(
                    f"[{sheet_name}!{cell_ref}] cell_formula_contains: セルが空です"
                )
            elif expected_text not in str(formula):
                errors.append(
                    f"[{sheet_name}!{cell_ref}] cell_formula_contains:"
                    f" '{expected_text}' が数式 '{formula}' に含まれていません"
                )
            else:
                print(f"  ✓ [{sheet_name}!{cell_ref}] 数式に '{expected_text}' を含む")

    wb_values.close()
    wb_formulas.close()
    return errors
