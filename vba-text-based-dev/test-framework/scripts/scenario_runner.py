"""シナリオ実行コアロジック

auto / manual 共通の VBA 実行・結果評価ロジックを提供する。

config.yaml の mode キーで実行モードを自動判定:
  - mode: manual → visible=True, testMode=False（ユーザーがダイアログを操作）
  - それ以外    → visible=False, testMode=True（自動実行）

環境変数:
  TOOL_TEST_ROOT: テスト対象ツールの test/ ディレクトリへの絶対パス。
                  tool_config.yaml・auto/・manual/・temp_dir/ の起点となる。
"""
import gc
import os
import re
import shutil
import time
from pathlib import Path

import yaml

DIALOG_LOG_FILENAME = "dialog_log.txt"

import openpyxl
import pywintypes
import xlwings as xw

from helpers.config_loader import load_scenario_config, validate_step
from helpers.xlsx_diff import compare_workbooks


TOOL_TEST_ROOT = Path(os.environ["TOOL_TEST_ROOT"])
TEMP_DIR = TOOL_TEST_ROOT / "temp_dir"

# tool_config.yaml からツール固有設定を読み込む
def _load_tool_config() -> dict:
    cfg_path = TOOL_TEST_ROOT / "tool_config.yaml"
    with open(cfg_path, encoding="utf-8") as f:
        return yaml.safe_load(f)

_tool_cfg = _load_tool_config()
XLSM_PATH = (TOOL_TEST_ROOT / _tool_cfg["xlsm_path"]).resolve()
XLSM_NAME = _tool_cfg["xlsm_name"]


def run_scenario(scenario_src_dir: Path, adapter=None) -> tuple[Path, dict]:
    """シナリオを実行し、(作業ディレクトリ, config) を返す。

    adapter が指定された場合は ExecutionOrchestrator 経由の新しいパスを使用する。
    adapter が None の場合は _legacy_run_scenario() を呼び、従来動作を維持する。

    config.yaml の mode が "manual" の場合は visible=True / testMode=False、
    それ以外は visible=False / testMode=True で実行する。

    Args:
        scenario_src_dir: シナリオのソースディレクトリ（auto/ or manual/ 配下）
        adapter: ToolAdapter 実装（None の場合はレガシーパスを使用）

    Returns:
        tuple: (work_dir, config)
    """
    if adapter is None:
        return _legacy_run_scenario(scenario_src_dir)

    from adapters.execution_orchestrator import get_orchestrator
    return get_orchestrator(XLSM_PATH, XLSM_NAME, TEMP_DIR).run_scenario(scenario_src_dir, adapter)


def _legacy_run_scenario(scenario_src_dir: Path) -> tuple[Path, dict]:
    """旧来のシナリオ実行（doctool 固有実装・後方互換）。

    adapter 引数なしで run_scenario() を呼んだ場合に使用される。
    既存の全シナリオとの互換性を保つため、このメソッドの内容は変更しない。
    """
    config = load_scenario_config(str(scenario_src_dir))
    is_manual = config.get("mode") == "manual"
    visible = is_manual
    test_mode = not is_manual
    scenario_name = scenario_src_dir.name

    # 作業ディレクトリ: temp_dir/scenarioXX/（実行のたびにクリア）
    work_dir = TEMP_DIR / scenario_name
    if work_dir.exists():
        shutil.rmtree(work_dir)
    work_dir.mkdir(parents=True)

    # シナリオのファイルを作業ディレクトリへコピー
    shutil.copytree(str(scenario_src_dir), str(work_dir), dirs_exist_ok=True)

    # xlsm をコピー
    xlsm_dest = work_dir / XLSM_NAME
    shutil.copy2(XLSM_PATH, xlsm_dest)

    # skip_open_files に該当しないファイルだけ Excel で開く
    skip_patterns = config.get("skip_open_files", [])
    all_input_files = sorted(
        f for f in work_dir.glob("*.xlsx")
        if "_expected" not in f.name
    )
    open_files = [
        f for f in all_input_files
        if not any(re.fullmatch(pat, f.name) for pat in skip_patterns)
    ]
    skipped = [f.name for f in all_input_files if f not in open_files]
    if skipped:
        print(f"[{scenario_name}] skip_open_files により除外: {skipped}")

    _execute_vba(scenario_name, xlsm_dest, open_files, config, visible, test_mode, work_dir)
    return work_dir, config


def evaluate_template_assertions(
    work_dir: Path,
    assertions: list[dict],
) -> list[str]:
    """テンプレートシートのアサーションを評価し、エラーメッセージのリストを返す。

    VBA 実行・保存後の xlsm ファイルを openpyxl で読み込み、以下を検証する:
    - category_count: 行7 B列以降の非空セル数（カテゴリ列数）
    - cell_formula_contains: 指定セルの数式が特定文字列を含む

    Args:
        work_dir: VBA 実行後の xlsm が格納されたディレクトリ
        assertions: config.yaml の template_assertions リスト

    Returns:
        list[str]: エラーメッセージのリスト（空なら全アサーション通過）
    """
    xlsm_path = work_dir / XLSM_NAME
    errors = []

    try:
        wb_values = openpyxl.load_workbook(str(xlsm_path), keep_vba=True, data_only=True)
        wb_formulas = openpyxl.load_workbook(str(xlsm_path), keep_vba=True, data_only=False)
    except Exception as e:
        return [f"template_assertions: xlsm の読み込みに失敗しました: {e}"]

    for assertion in assertions:
        sheet_name = assertion.get("sheet", "レビュー結果シートテンプレート")

        if sheet_name not in wb_values.sheetnames:
            errors.append(f"template_assertions: シート '{sheet_name}' が存在しません")
            continue

        ws_v = wb_values[sheet_name]
        ws_f = wb_formulas[sheet_name]

        # カテゴリ列数チェック（行7 B列以降の非空セル数）
        if "category_count" in assertion:
            expected = assertion["category_count"]
            count = 0
            col = 2  # B列
            while ws_v.cell(7, col).value is not None:
                count += 1
                col += 1
            if count != expected:
                errors.append(
                    f"[{sheet_name}] category_count: expected={expected}, actual={count}"
                )
            else:
                print(f"  ✓ [{sheet_name}] category_count = {count}")

        # 数式の文字列チェック
        for fc in assertion.get("cell_formula_contains", []):
            cell_ref = fc["cell"]
            expected_text = fc["contains"]
            formula = ws_f[cell_ref].value
            if formula is None:
                errors.append(
                    f"[{sheet_name}!{cell_ref}] cell_formula_contains: セルが空です"
                )
            elif expected_text not in str(formula):
                errors.append(
                    f"[{sheet_name}!{cell_ref}] cell_formula_contains:"
                    f" '{expected_text}' が数式 '{formula}' に含まれていません"
                )
            else:
                print(f"  ✓ [{sheet_name}!{cell_ref}] 数式に '{expected_text}' を含む")

    wb_values.close()
    wb_formulas.close()
    return errors


def evaluate_scenario(work_dir: Path, scenario_src_dir: Path, config: dict, adapter=None) -> list[str]:
    """シナリオ結果を評価し、エラーメッセージのリストを返す。

    adapter が指定された場合は ExecutionOrchestrator 経由の新しいパスを使用する。
    adapter が None の場合は従来の doctool 固有実装を維持する。

    評価内容:
      1. file_expectations に指定されたファイル: assert_no_sheets などを評価
      2. それ以外のファイル: _expected.xlsx が存在すれば全シート Gold Master 比較

    Args:
        work_dir: VBA 実行後の actual ファイルが格納されたディレクトリ
        scenario_src_dir: シナリオのソースディレクトリ（_expected.xlsx の格納元）
        config: load_scenario_config で読み込んだ設定 dict

    Returns:
        list[str]: エラーメッセージのリスト（空なら全アサーション通過）
    """
    if adapter is not None:
        from adapters.execution_orchestrator import get_orchestrator
        return get_orchestrator(XLSM_PATH, XLSM_NAME, TEMP_DIR).evaluate_scenario(
            work_dir, scenario_src_dir, config, adapter
        )

    expectations = config.get("file_expectations", [])
    excluded_cells = config.get("excluded_cells", [])
    errors = []
    evaluated_files = set()
    assertion_count = 0  # 実施したアサーションの数（0件はテスト失敗）

    # file_expectations で指定されたファイルを処理
    for expectation in expectations:
        pattern = expectation["pattern"]
        matched = sorted(
            f for f in work_dir.glob("*.xlsx")
            if re.search(pattern, f.name) and "_expected" not in f.name
        )
        if not matched:
            errors.append(f"パターン '{pattern}' にマッチするファイルが見つかりません")
            continue

        actual_path = matched[0]
        evaluated_files.add(actual_path.name)

        # assert_no_sheets: 指定シートが存在しないことを確認
        # 空リスト（assert_no_sheets: []）は意図的な「評価対象外」のためカウントしない
        for sheet_name in expectation.get("assert_no_sheets", []):
            wb = openpyxl.load_workbook(str(actual_path), data_only=True)
            if sheet_name in wb.sheetnames:
                errors.append(
                    f"[{actual_path.name}] シート '{sheet_name}' が"
                    f"存在してはいけませんが存在します"
                )
            else:
                print(f"  ✓ [{actual_path.name}] '{sheet_name}' が存在しない - OK")
            wb.close()
            assertion_count += 1

    # デフォルト: _expected.xlsx が存在する全ファイルを全シート Gold Master 比較
    all_input = sorted(
        f for f in work_dir.glob("*.xlsx")
        if "_expected" not in f.name
    )
    for actual_path in all_input:
        if actual_path.name in evaluated_files:
            continue  # file_expectations で処理済み

        expected_path = scenario_src_dir / f"{actual_path.stem}_expected{actual_path.suffix}"
        if not expected_path.exists():
            print(f"  - [{actual_path.name}] Gold Master なし → スキップ")
            continue

        result = compare_workbooks(
            str(actual_path),
            str(expected_path),
            excluded_cells=excluded_cells,
            compare_fill=True,
            compare_col_widths=True,
            compare_print_area=True,
            compare_borders=True,
            max_diffs=20,
        )
        if result.matches:
            print(f"  ✓ [{actual_path.name}] Gold Master 比較 - OK")
        else:
            errors.append(
                f"[{actual_path.name}] Gold Master 比較 FAILED:\n" +
                "\n".join(f"  {d}" for d in result.diffs)
            )
        assertion_count += 1

    # template_assertions の評価
    template_assertions = config.get("template_assertions", [])
    if template_assertions:
        template_errors = evaluate_template_assertions(work_dir, template_assertions)
        errors.extend(template_errors)
        assertion_count += len(template_assertions)

    # expected_messages の評価（双方向チェック）
    # config に expected_messages が定義されている場合のみ評価する（None = 未定義はスキップ）
    # - 正方向: expected に列挙したメッセージIDが全て dialog_log に出現しているか
    # - 逆方向: dialog_log に出現したメッセージIDが expected 以外のものを含んでいないか
    expected_messages = config.get("expected_messages")
    if expected_messages is not None:
        log_path = work_dir / DIALOG_LOG_FILENAME
        if not log_path.exists():
            if expected_messages:
                errors.append(
                    "expected_messages: ダイアログログ（dialog_log.txt）が存在しません。"
                    " testMode で extract ステップが実行されていない可能性があります。"
                )
            # expected_messages: [] でログなしは正常（ダイアログが出ないシナリオ）
        else:
            log_content = log_path.read_text(encoding="utf-8")
            found_ids = re.findall(r'\[MSG:([A-Z]\d+)\]', log_content)
            found_set = set(found_ids)
            expected_set = set(expected_messages)

            # 正方向: expected に列挙した全IDがログに存在するか
            for expected_id in expected_messages:
                if expected_id in found_set:
                    print(f"  ✓ expected_messages: '{expected_id}' を確認")
                else:
                    errors.append(
                        f"expected_messages: '{expected_id}' がダイアログログに見つかりません"
                        f" (実際のログ: {sorted(found_set)})"
                    )

            # 逆方向: ログに想定外のIDが含まれていないか
            unexpected = found_set - expected_set
            if unexpected:
                errors.append(
                    f"expected_messages: 想定外のメッセージが出力されました: {sorted(unexpected)}"
                    f" (expected: {sorted(expected_set)})"
                )
            else:
                print(f"  ✓ expected_messages: 想定外メッセージなし")

        # expected_messages は assertion_count に含めない。
        # シート有無確認か Gold Master 比較のいずれかが必ず存在することを別途保証する。

    # 構造的アサーション（シート有無確認 / Gold Master 比較 / template_assertions）が
    # 1件もない場合はテスト設定ミスとして失敗。
    # ただし file_expectations: [] が明示指定されている場合は「出力なし」が意図的であるため除外。
    # （例: 削除対象なし通知のシナリオ ― E09/E10 発火のみを expected_messages で検証）
    if assertion_count == 0:
        if "file_expectations" in config and config["file_expectations"] == []:
            pass  # 出力ファイルなしが意図的なシナリオ（E09/E10 等）
        else:
            errors.append(
                "構造的アサーションが1件もありません。"
                "_expected.xlsx を配置するか、file_expectations で assert_no_sheets を設定するか、"
                "template_assertions を設定してください。"
                "（expected_messages はダイアログ検証専用のため、この条件を満たしません）"
                "出力ファイルが存在しない意図的なシナリオの場合は file_expectations: [] を設定してください。"
            )

    return errors


# ============================================================
# 内部ヘルパー
# ============================================================

def _execute_vba(
    scenario_name: str,
    xlsm_dest: Path,
    open_files: list[Path],
    config: dict,
    visible: bool,
    test_mode: bool,
    work_dir: Path,
) -> None:
    """VBA マクロを実行する。"""
    app = None
    open_wbs = []
    xlsm_wb = None

    try:
        print(f"[{scenario_name}] Excel を起動中... (visible={visible})")
        app = xw.App(visible=visible)

        # Excel 起動時に自動作成されるデフォルトブック（Book1 等）を閉じる
        for book in list(app.books):
            try:
                book.close()
            except Exception:
                pass

        for f in open_files:
            print(f"[{scenario_name}] ファイルを開く: {f.name}")
            wb = app.books.open(str(f))
            open_wbs.append((f.name, wb))

        print(f"[{scenario_name}] xlsm を開く...")
        xlsm_wb = app.books.open(str(xlsm_dest))

        # setup キーによる xlsm 事前設定
        setup = config.get("setup", {})
        if setup:
            settings_ws = xlsm_wb.sheets["基本設定"]
            if "use_review_record" in setup:
                settings_ws["B2"].value = setup["use_review_record"]
                print(f"[{scenario_name}] setup: use_review_record={setup['use_review_record']}")
            if "use_summary" in setup:
                settings_ws["B3"].value = setup["use_summary"]
                print(f"[{scenario_name}] setup: use_summary={setup['use_summary']}")
            if "review_list_file" in setup:
                review_list_path = work_dir / setup["review_list_file"]
                xlsm_wb.names["REVIEW_LIST_FILEPATH"].refers_to_range.value = str(review_list_path)
                print(f"[{scenario_name}] setup: REVIEW_LIST_FILEPATH={review_list_path}")
            if "categories" in setup:
                _apply_categories(scenario_name, xlsm_wb, setup["categories"])
            if "named_ranges" in setup:
                for range_name, value in setup["named_ranges"].items():
                    try:
                        xlsm_wb.names[range_name].refers_to_range.value = value
                        print(f"[{scenario_name}] setup: named_ranges[{range_name}]={repr(value)}")
                    except Exception as e:
                        print(f"[{scenario_name}] setup: named_ranges warning [{range_name}]: {e}")
            if "item_mapping_cells" in setup:
                map_ws = xlsm_wb.sheets["項目マッピング設定"]
                for cell_ref, value in setup["item_mapping_cells"].items():
                    map_ws[cell_ref].value = value
                    print(f"[{scenario_name}] setup: item_mapping_cells[{cell_ref}]={repr(value)}")

        macro = xlsm_wb.macro("Sheet1.CmdGen_Click_Core")

        for step_idx, step in enumerate(config["steps"], 1):
            validate_step(step)
            action = step["action"]

            if action == "extract":
                # ステップ固有のカテゴリ設定（省略時はグローバル setup.categories を維持）
                if "categories" in step:
                    _apply_categories(scenario_name, xlsm_wb, step["categories"])

                review_times = step["review_times"]
                repeat = step.get("repeat", 1)
                print(
                    f"[{scenario_name}] Step {step_idx}: extract"
                    f" (REVIEW_TIMES={review_times}, repeat={repeat})"
                )
                for run_num in range(repeat):
                    xlsm_wb.names["REVIEW_TIMES"].refers_to_range.value = review_times
                    if test_mode:
                        xlsm_wb.macro("Module2.ClearDialogLog")()
                    print(f"[{scenario_name}]   マクロ実行中 (run {run_num + 1}/{repeat})...")
                    _run_macro_with_retry(scenario_name, macro, test_mode)
                    if test_mode:
                        _flush_dialog_log(scenario_name, xlsm_wb, work_dir)

            elif action == "delete_comments":
                print(f"[{scenario_name}] Step {step_idx}: delete_comments")
                _run_delete_macro(
                    scenario_name, xlsm_wb,
                    "Module2.DelAllReviewComments_Click_Core", test_mode, work_dir,
                )

            elif action == "delete_sheets":
                print(f"[{scenario_name}] Step {step_idx}: delete_sheets")
                _run_delete_macro(
                    scenario_name, xlsm_wb,
                    "Module2.DelAllReviewResultSheets_Click_Core", test_mode, work_dir,
                )

        print(f"[{scenario_name}] 全ステップ完了")

        print(f"[{scenario_name}] ファイルを保存中...")
        python_opened_names = {wb.name for _, wb in open_wbs} | {XLSM_NAME}
        for _, wb in open_wbs:
            try:
                wb.save()
            except Exception as e:
                print(f"[{scenario_name}]   保存警告: {e}")
        # VBA が自分で開いたファイル（skip_open_files で除外されたもの）も保存する
        for book in app.books:
            if book.name not in python_opened_names:
                try:
                    book.save()
                    print(f"[{scenario_name}]   VBA開きファイル保存: {book.name}")
                except Exception as e:
                    print(f"[{scenario_name}]   保存警告 (VBA-opened): {e}")
        xlsm_wb.save()
        print(f"[{scenario_name}] Saved → temp_dir/{scenario_name}/")

    except Exception as e:
        raise RuntimeError(f"[{scenario_name}] VBA 実行エラー: {e}") from e

    finally:
        _cleanup_excel(scenario_name, app, open_wbs, xlsm_wb)


def _apply_categories(
    scenario_name: str,
    xlsm_wb,
    categories: list[dict],
) -> None:
    """指摘分類マッピング設定シートをカテゴリリストで上書きする。

    Args:
        scenario_name: ログ用シナリオ名
        xlsm_wb: xlwings Workbook オブジェクト
        categories: [{"alias": "a", "name": "01_要件漏れ"}, ...] のリスト
    """
    cat_ws = xlsm_wb.sheets["指摘分類マッピング設定"]
    # 既存データを行2以降クリア
    last_row = cat_ws.range("A1").current_region.last_cell.row
    if last_row >= 2:
        cat_ws.range(f"A2:B{last_row}").clear_contents()
    # 新しいカテゴリを書き込む
    for i, cat in enumerate(categories, start=2):
        cat_ws.range(f"A{i}").value = cat["alias"]
        cat_ws.range(f"B{i}").value = cat["name"]
    print(f"[{scenario_name}] categories: {len(categories)} カテゴリを設定 ({[c['alias'] for c in categories]})")


def _flush_dialog_log(scenario_name: str, xlsm_wb, work_dir: Path) -> None:
    """VBA ダイアログログを dialog_log.txt に追記し、コンソールにも出力する。"""
    dialog_log = xlsm_wb.macro("Module2.GetDialogLog")()
    if dialog_log:
        log_path = work_dir / DIALOG_LOG_FILENAME
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(dialog_log + "\n")
        for line in dialog_log.split("\n"):
            if line:
                print(f"[{scenario_name}]   [LOG] {line}")


def _run_macro_with_retry(scenario_name: str, macro, test_mode: bool, max_retries: int = 3) -> None:
    """VBAマクロをCOMエラー時にリトライ付きで実行する。

    エラーコード 0x80010100 (RPC_E_SERVERCALL_ISBUSY) は Excel が COM メッセージポンプに
    応答できない状態を示す。一時的なビジー状態のため、待機後に再試行することで回復できる。
    """
    # COM_E_SERVER_ISBUSY: Excelがビジー状態でCOM呼び出しに応答できない
    COM_E_SERVER_ISBUSY = -2147417856  # 0x80010100
    retry_wait_seconds = 4

    for attempt in range(1, max_retries + 2):
        try:
            macro(test_mode)
            return
        except pywintypes.com_error as e:
            hresult = e.hresult if hasattr(e, "hresult") else (e.args[0] if e.args else None)
            if hresult == COM_E_SERVER_ISBUSY and attempt <= max_retries:
                print(
                    f"[{scenario_name}]   COMエラー (0x80010100: サーバービジー)。"
                    f"{retry_wait_seconds}秒後にリトライ ({attempt}/{max_retries})..."
                )
                time.sleep(retry_wait_seconds)
            else:
                raise


def _run_delete_macro(
    scenario_name: str,
    xlsm_wb,
    macro_path: str,
    test_mode: bool,
    work_dir: Path,
) -> None:
    """削除系マクロを実行する。"""
    try:
        xlsm_wb.macro("Module2.ClearDialogLog")()
        xlsm_wb.macro(macro_path)(test_mode)
        _flush_dialog_log(scenario_name, xlsm_wb, work_dir)
    except Exception as e:
        print(f"[{scenario_name}]   Warning: {e}")


def _cleanup_excel(
    scenario_name: str,
    app,
    open_wbs: list,
    xlsm_wb,
) -> None:
    """Excel プロセスをクリーンアップする。

    COM オブジェクトへの参照を明示的に None で破棄してから app.quit() を呼ぶ。
    参照を残したまま quit すると GC 時に死んだサーバーへ接続しようとして
    'Windows fatal exception: code 0x800706ba' (RPC_S_SERVER_UNAVAILABLE) が発生する。
    """
    for _, wb in open_wbs:
        try:
            wb.close()
        except Exception:
            pass
    open_wbs.clear()

    if xlsm_wb:
        try:
            xlsm_wb.close()
        except Exception:
            pass
        xlsm_wb = None  # noqa: F841 — COM参照を明示破棄

    gc.collect()  # close後・quit前にCOM参照を解放

    if app:
        try:
            app.quit()
        except Exception:
            pass
        app = None  # noqa: F841

    gc.collect()

    try:
        import psutil
        for i in range(10):
            if not any(
                p.name().lower() == "excel.exe"
                for p in psutil.process_iter(["name"])
            ):
                print(f"[{scenario_name}] Excel 終了確認 ({i + 1}s)")
                break
            time.sleep(1)
        else:
            print(f"[{scenario_name}] 警告: Excel が 10秒後も残留中")
    except Exception:
        pass
