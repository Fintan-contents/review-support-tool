"""Excel アサーションヘルパーモジュール

VBA出力xlsxの検証に便利な関数を提供する。
"""
import openpyxl
from typing import Optional


def get_result_sheet(wb: openpyxl.Workbook, review_times: int = 1) -> openpyxl.worksheet.worksheet.Worksheet:
    """レビュー結果N回目シートを取得
    
    Args:
        wb: Excelワークブック
        review_times: レビュー回数（1, 2, 3...）
        
    Returns:
        レビュー結果シート
        
    Raises:
        AssertionError: シートが存在しない場合
    """
    name = f"レビュー結果{review_times}回目"
    assert name in wb.sheetnames, \
        f"Sheet '{name}' not found. Available sheets: {wb.sheetnames}"
    return wb[name]


def get_detail_rows(ws: openpyxl.worksheet.worksheet.Worksheet, start_row: int = 15) -> list[dict]:
    """指摘詳細行を辞書リストとして取得
    
    Args:
        ws: ワークシート
        start_row: 指摘詳細の開始行（デフォルト15）
        
    Returns:
        指摘詳細の辞書リスト。各辞書は以下のキーを含む：
        - row: 行番号
        - sheet_name: シート名
        - position: セル位置（例: "B5"）
        - reviewer: レビュー者名
        - category: カテゴリ（a, b, c等）
        - comment: 指摘内容
        - fix_status: 対応状況（"済", "未"等）
    """
    # VBA定数と対応（DETAIL_COL_*）
    DETAIL_COL_SHEET = 1
    DETAIL_COL_POSITION = 2
    DETAIL_COL_REVIEWER = 3
    DETAIL_COL_CATEGORY = 4
    DETAIL_COL_REVIEW_COMMENT = 5
    DETAIL_COL_FIX_STATUS = 11
    
    rows = []
    for row_idx in range(start_row, ws.max_row + 1):
        position = ws.cell(row_idx, DETAIL_COL_POSITION).value
        # 場所（セル位置）が空の場合はスキップ
        if position:
            rows.append({
                "row": row_idx,
                "sheet_name": ws.cell(row_idx, DETAIL_COL_SHEET).value,
                "position": str(position),
                "reviewer": ws.cell(row_idx, DETAIL_COL_REVIEWER).value,
                "category": ws.cell(row_idx, DETAIL_COL_CATEGORY).value,
                "comment": ws.cell(row_idx, DETAIL_COL_REVIEW_COMMENT).value,
                "fix_status": ws.cell(row_idx, DETAIL_COL_FIX_STATUS).value,
            })
    return rows


def has_error_sheet(wb: openpyxl.Workbook) -> bool:
    """エラーシートが存在するか
    
    Args:
        wb: Excelワークブック
        
    Returns:
        エラーシートが存在する場合True
    """
    return "エラーシート" in wb.sheetnames


def count_review_record_issues(
    wb: openpyxl.Workbook,
    sheet_name: str = "レビュー指摘一覧",
    start_row: int = 5,
    review_count_col: int = 2
) -> int:
    """レビュー記録票の指摘件数をカウント
    
    Args:
        wb: レビュー記録票ワークブック
        sheet_name: 指摘一覧シート名
        start_row: 指摘一覧の開始行
        review_count_col: レビュー回数列（B列=2）
        
    Returns:
        指摘件数
    """
    if sheet_name not in wb.sheetnames:
        return 0
    
    ws = wb[sheet_name]
    count = 0
    
    for row_idx in range(start_row, ws.max_row + 1):
        # A列にENDマーカーがあれば終了
        marker = ws.cell(row_idx, 1).value
        if marker and str(marker).strip().upper() == "END":
            break
        
        # レビュー回数列に値があれば有効行としてカウント
        if ws.cell(row_idx, review_count_col).value is not None:
            count += 1
    
    return count


def get_header_info(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict:
    """レビュー結果シートのヘッダー情報を取得
    
    Args:
        ws: レビュー結果シート
        
    Returns:
        ヘッダー情報の辞書：
        - target_file: 対象ファイル名（A4）
        - review_count: レビュー回数（C4）
        - review_date: 実施日（E4）
        - start_time: 開始時刻（F4）
        - end_time: 終了時刻（H4）
        - review_time: レビュー時間（I4）
    """
    # VBA定数と対応
    COL_TARGET = 1;  ROW_TARGET = 4
    COL_COUNT = 3;   ROW_COUNT = 4
    COL_DATE = 5;    ROW_DATE = 4
    COL_START = 6;   ROW_START = 4
    COL_END = 8;     ROW_END = 4
    COL_BREAK = 9;   ROW_BREAK = 4
    
    return {
        "target_file": ws.cell(ROW_TARGET, COL_TARGET).value,
        "review_count": ws.cell(ROW_COUNT, COL_COUNT).value,
        "review_date": ws.cell(ROW_DATE, COL_DATE).value,
        "start_time": ws.cell(ROW_START, COL_START).value,
        "end_time": ws.cell(ROW_END, COL_END).value,
        "review_time": ws.cell(ROW_BREAK, COL_BREAK).value,
    }


def get_category_headers(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict:
    """カテゴリヘッダー行を取得
    
    Args:
        ws: レビュー結果シート
        
    Returns:
        カテゴリ情報の辞書：
        - abbreviations: カテゴリ略称のリスト(7行目: a, b, c, ...)
        - names: カテゴリ名のリスト(8行目: 01_要件漏れ, ...)
    """
    ROW_CATEGORY_ABBR = 7  # カテゴリ略称（a～i）
    ROW_CATEGORY_NAME = 8  # カテゴリ名（01_要件漏れ等）
    COL_CATEGORY_START = 2
    
    abbreviations = []
    names = []
    
    # 最大9カテゴリを想定（a-i）
    for col in range(COL_CATEGORY_START, COL_CATEGORY_START + 10):
        abbr = ws.cell(ROW_CATEGORY_ABBR, col).value
        name = ws.cell(ROW_CATEGORY_NAME, col).value
        
        if abbr and str(abbr).strip():
            abbreviations.append(str(abbr).strip())
        if name and str(name).strip():
            names.append(str(name).strip())
    
    return {
        "abbreviations": abbreviations,
        "names": names
    }


def get_category_counts(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict:
    """カテゴリ別集計を取得
    
    Args:
        ws: レビュー結果シート
        
    Returns:
        集計情報の辞書：
        - fixed: 「済」の件数辞書 {カテゴリ略称: 件数}
        - unfixed: 「未」の件数辞書 {カテゴリ略称: 件数}
        - fixed_total: 「済」の合計件数
        - unfixed_total: 「未」の合計件数
    """
    ROW_FIXED = 9   # 「済」集計行
    ROW_UNFIXED = 10  # 「未」集計行
    COL_CATEGORY_START = 2
    COL_TOTAL = 11  # 合計列
    
    # カテゴリ略称を取得
    categories = get_category_headers(ws)["abbreviations"]
    
    fixed = {}
    unfixed = {}
    
    for idx, cat in enumerate(categories):
        col = COL_CATEGORY_START + idx
        fixed_val = ws.cell(ROW_FIXED, col).value
        unfixed_val = ws.cell(ROW_UNFIXED, col).value
        
        if fixed_val:
            fixed[cat] = int(fixed_val)
        if unfixed_val:
            unfixed[cat] = int(unfixed_val)
    
    return {
        "fixed": fixed,
        "unfixed": unfixed,
        "fixed_total": ws.cell(ROW_FIXED, COL_TOTAL).value or 0,
        "unfixed_total": ws.cell(ROW_UNFIXED, COL_TOTAL).value or 0,
    }


def get_review_record_issues(
    wb: openpyxl.Workbook,
    sheet_name: str = "レビュー指摘一覧",
    start_row: int = 5
) -> list[dict]:
    """レビュー記録票の指摘一覧を取得
    
    Args:
        wb: レビュー記録票ワークブック
        sheet_name: 指摘一覧シート名
        start_row: 指摘一覧の開始行
        
    Returns:
        指摘詳細の辞書リスト。各辞書は以下のキーを含む：
        - row: 行番号
        - review_count: レビュー回数（B列）
        - position: セル位置（C列）
        - category: カテゴリ（D列）
        - reviewer: レビュー者（E列）
        - review_date: レビュー日（F列）
        - fix_status: 対応状況（G列）
        - issue_comment: 指摘内容（H列）- デリミタ前
        - response_comment: 対応内容（J列）- デリミタ後
    """
    if sheet_name not in wb.sheetnames:
        return []
    
    ws = wb[sheet_name]
    issues = []
    
    # 列定義（VBA: Module1.bas の RECORD_COL_* 定数と対応）
    COL_REVIEW_COUNT = 2  # B列
    COL_POSITION = 3      # C列
    COL_CATEGORY = 4      # D列
    COL_REVIEWER = 5      # E列
    COL_REVIEW_DATE = 6   # F列
    COL_FIX_STATUS = 7    # G列
    COL_ISSUE_COMMENT = 8 # H列（デリミタ前）
    COL_RESPONSE_COMMENT = 10  # J列（デリミタ後）
    
    for row_idx in range(start_row, ws.max_row + 1):
        # A列にENDマーカーがあれば終了
        marker = ws.cell(row_idx, 1).value
        if marker and str(marker).strip().upper() == "END":
            break
        
        # レビュー回数列に値があれば有効行
        review_count = ws.cell(row_idx, COL_REVIEW_COUNT).value
        if review_count is not None:
            issues.append({
                "row": row_idx,
                "review_count": review_count,
                "position": ws.cell(row_idx, COL_POSITION).value,
                "category": ws.cell(row_idx, COL_CATEGORY).value,
                "reviewer": ws.cell(row_idx, COL_REVIEWER).value,
                "review_date": ws.cell(row_idx, COL_REVIEW_DATE).value,
                "fix_status": ws.cell(row_idx, COL_FIX_STATUS).value,
                "issue_comment": ws.cell(row_idx, COL_ISSUE_COMMENT).value,
                "response_comment": ws.cell(row_idx, COL_RESPONSE_COMMENT).value,
            })
    
    return issues

