"""
ベンチマーク用フィクスチャ生成スクリプト

指定したカテゴリ数のテストフィクスチャ（xlsx）を生成する。
S29 のフィクスチャをテンプレートとして使用し、指定カテゴリ数分のコメントを
複数シートに均等分散して挿入する。

使用方法:
    python create_fixture.py --categories 50 --output-dir ./tmp/bench_50
    python create_fixture.py --categories 702
"""

import argparse
import shutil
import string
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment


TEST_DIR = Path(__file__).parent.parent  # doctool/test
S29_DIR = TEST_DIR / "auto" / "scenario29"

DESIGN_DOC_TEMPLATE = S29_DIR / "システム機能設計書_サンプル_S29.xlsx"
REVIEW_RECORD_TEMPLATE = S29_DIR / "システム機能設計書_サンプル_S29_レビュー記録票.xlsx"
REVIEW_SUMMARY_TEMPLATE = S29_DIR / "レビュー記録サマリ_S29.xlsx"

MAX_CATEGORIES = 702


def gen_aliases(n: int) -> list[str]:
    """n 個のカテゴリエイリアスを生成する。

    1-26: a-z
    27-702: aa, ab, ..., zz
    """
    if n > MAX_CATEGORIES:
        raise ValueError(f"カテゴリ数は {MAX_CATEGORIES} 以下にしてください（指定値: {n}）")
    aliases: list[str] = []
    # 1文字: a-z
    for c in string.ascii_lowercase:
        aliases.append(c)
        if len(aliases) >= n:
            return aliases
    # 2文字: aa, ab, ..., zz
    for first in string.ascii_lowercase:
        for second in string.ascii_lowercase:
            aliases.append(first + second)
            if len(aliases) >= n:
                return aliases
    return aliases


def add_comment(ws, cell_ref: str, text: str, author: str = "山田　太郎") -> None:
    ws[cell_ref].comment = Comment(text, author)


def remove_existing_comments(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment:
                cell.comment = None


def create_fixture(categories: int, output_dir: Path) -> Path:
    """フィクスチャを生成し、設計書 xlsx のパスを返す。

    Args:
        categories: カテゴリ数
        output_dir: 出力先ディレクトリ（存在しない場合は作成）

    Returns:
        生成した設計書 xlsx のパス
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    all_cats = gen_aliases(categories)

    # ---- 設計書フィクスチャ ----
    dst_design = output_dir / "設計書_bench.xlsx"
    wb = load_workbook(DESIGN_DOC_TEMPLATE)

    excluded = {"表紙", "変更履歴", "目次", "データ"}
    target_sheets = [s for s in wb.sheetnames if s not in excluded]
    if not target_sheets:
        raise RuntimeError("コメント挿入可能なシートが存在しません")

    for sn in target_sheets:
        remove_existing_comments(wb[sn])

    # 日時コメントをシート1に挿入
    ws0 = wb[target_sheets[0]]
    add_comment(
        ws0, "A4",
        "山田　太郎:*\n実施日:2023/08/01\n開始:09:00\n終了:11:00\nレビュー時間:120",
    )

    # カテゴリコメントを均等分散
    num_sheets = len(target_sheets)
    base, rem = divmod(len(all_cats), num_sheets)
    start = 0
    for i, sheet_name in enumerate(target_sheets):
        count = base + (1 if i < rem else 0)
        sheet_cats = all_cats[start: start + count]
        start += count
        ws = wb[sheet_name]
        for idx, alias in enumerate(sheet_cats):
            row = 5 + idx
            content = f"{alias}カテゴリの指摘内容サンプル。"
            if idx % 10 == 0:
                text = f"山田　太郎:{alias}\n{content}\n---\n対応済みです。\n済"
            else:
                text = f"山田　太郎:{alias}\n{content}"
            add_comment(ws, f"B{row}", text)

    wb.save(dst_design)
    print(f"  設計書: {dst_design} ({categories} カテゴリ)")

    # ---- レビュー記録票・サマリ コピー ----
    dst_record = output_dir / "設計書_bench_レビュー記録票.xlsx"
    shutil.copy2(REVIEW_RECORD_TEMPLATE, dst_record)

    dst_summary = output_dir / "レビュー記録サマリ_bench.xlsx"
    shutil.copy2(REVIEW_SUMMARY_TEMPLATE, dst_summary)

    return dst_design


def main() -> None:
    parser = argparse.ArgumentParser(
        description="ベンチマーク用テストフィクスチャを生成する"
    )
    parser.add_argument(
        "--categories", "-c",
        type=int,
        required=True,
        help=f"カテゴリ数（1〜{MAX_CATEGORIES}）",
    )
    parser.add_argument(
        "--output-dir", "-o",
        type=Path,
        default=None,
        help="出力先ディレクトリ（省略時: ./tmp/bench_<categories>）",
    )
    args = parser.parse_args()

    if args.output_dir is None:
        args.output_dir = Path(__file__).parent / "tmp" / f"bench_{args.categories}"

    print(f"カテゴリ数: {args.categories}")
    print(f"出力先: {args.output_dir}")
    create_fixture(args.categories, args.output_dir)
    print("完了")


if __name__ == "__main__":
    main()
