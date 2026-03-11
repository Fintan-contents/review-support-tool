"""
xlsx_diff モジュールのユニットテスト

compare_cells, _build_excluded_set, compare_workbooks, print_diff_report の
正常系・異常系を検証する。
"""
import io
import sys
from datetime import datetime

import openpyxl
import pytest

from helpers.xlsx_diff import (
    DiffResult,
    _build_excluded_set,
    compare_cells,
    compare_workbooks,
    print_diff_report,
)


# ---------------------------------------------------------------------------
# compare_cells
# ---------------------------------------------------------------------------

class TestCompareCells:
    """compare_cells のテスト"""

    def test_both_none(self):
        """正常系: 両方 None の場合は True を返す"""
        assert compare_cells(None, None) is True

    def test_actual_none(self):
        """異常系: actual のみ None の場合は False を返す"""
        assert compare_cells(None, "value") is False

    def test_expected_none(self):
        """異常系: expected のみ None の場合は False を返す"""
        assert compare_cells("value", None) is False

    def test_equal_strings(self):
        """正常系: 同一文字列は True を返す"""
        assert compare_cells("text", "text") is True

    def test_string_with_whitespace(self):
        """正常系: 前後の空白を除いた文字列が一致する場合は True を返す"""
        assert compare_cells("  text  ", "text") is True

    def test_different_strings(self):
        """異常系: 異なる文字列は False を返す"""
        assert compare_cells("abc", "def") is False

    def test_equal_integers(self):
        """正常系: 同一の整数値は True を返す"""
        assert compare_cells(42, 42) is True

    def test_equal_float_precision(self):
        """正常系: 浮動小数点誤差の範囲内（< 1e-9）は True を返す"""
        assert compare_cells(1.0, 1.0 + 1e-10) is True

    def test_different_numbers(self):
        """異常系: 差が 1e-9 を超える数値は False を返す"""
        assert compare_cells(1.0, 2.0) is False

    def test_datetime_within_tolerance(self):
        """正常系: 1 秒未満の差異の datetime は True を返す"""
        dt1 = datetime(2024, 1, 1, 12, 0, 0)
        dt2 = datetime(2024, 1, 1, 12, 0, 0, 500000)  # 0.5 秒差
        assert compare_cells(dt1, dt2) is True

    def test_datetime_exceeds_tolerance(self):
        """異常系: 1 秒以上の差異の datetime は False を返す"""
        dt1 = datetime(2024, 1, 1, 12, 0, 0)
        dt2 = datetime(2024, 1, 1, 12, 0, 2)  # 2 秒差
        assert compare_cells(dt1, dt2) is False


# ---------------------------------------------------------------------------
# _build_excluded_set
# ---------------------------------------------------------------------------

class TestBuildExcludedSet:
    """_build_excluded_set のテスト"""

    def test_empty_input(self):
        """正常系: 空リストを渡した場合は空辞書を返す"""
        assert _build_excluded_set([]) == {}

    def test_none_input(self):
        """正常系: None を渡した場合は空辞書を返す"""
        assert _build_excluded_set(None) == {}

    def test_single_entry(self):
        """正常系: 1 エントリを正しく変換する"""
        excluded = [{"sheet": "Sheet1", "cells": ["A1", "B2"]}]
        result = _build_excluded_set(excluded)
        assert result == {"Sheet1": {"A1", "B2"}}

    def test_multiple_entries_same_sheet(self):
        """正常系: 同一シートの複数エントリをマージする"""
        excluded = [
            {"sheet": "Sheet1", "cells": ["A1"]},
            {"sheet": "Sheet1", "cells": ["B2"]},
        ]
        result = _build_excluded_set(excluded)
        assert result == {"Sheet1": {"A1", "B2"}}

    def test_multiple_sheets(self):
        """正常系: 複数シートのエントリを個別に保持する"""
        excluded = [
            {"sheet": "Sheet1", "cells": ["A1"]},
            {"sheet": "Sheet2", "cells": ["C3"]},
        ]
        result = _build_excluded_set(excluded)
        assert result == {"Sheet1": {"A1"}, "Sheet2": {"C3"}}


# ---------------------------------------------------------------------------
# compare_workbooks（xlsxファイルを使用）
# ---------------------------------------------------------------------------

def _make_workbook(tmp_path, filename: str, data: dict) -> str:
    """テスト用の xlsx ファイルを作成するユーティリティ"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, cells in data.items():
        ws = wb.create_sheet(title=sheet_name)
        for cell_ref, value in cells.items():
            ws[cell_ref] = value
    path = str(tmp_path / filename)
    wb.save(path)
    return path


class TestCompareWorkbooks:
    """compare_workbooks のテスト"""

    def test_identical_workbooks(self, tmp_path):
        """正常系: 内容が完全に一致するワークブックは matches=True を返す"""
        data = {"Sheet1": {"A1": "value", "B2": 100}}
        actual = _make_workbook(tmp_path, "actual.xlsx", data)
        expected = _make_workbook(tmp_path, "expected.xlsx", data)

        result = compare_workbooks(actual, expected)

        assert result.matches is True
        assert result.diffs == []

    def test_cell_value_mismatch(self, tmp_path):
        """異常系: セル値が異なる場合は matches=False かつ差分メッセージを返す"""
        actual = _make_workbook(tmp_path, "actual.xlsx", {"Sheet1": {"A1": "wrong"}})
        expected = _make_workbook(tmp_path, "expected.xlsx", {"Sheet1": {"A1": "correct"}})

        result = compare_workbooks(actual, expected)

        assert result.matches is False
        assert any("A1" in d for d in result.diffs)

    def test_missing_sheet_in_actual(self, tmp_path):
        """異常系: actual に期待シートが存在しない場合は差分メッセージを返す"""
        actual = _make_workbook(tmp_path, "actual.xlsx", {"OtherSheet": {"A1": "x"}})
        expected = _make_workbook(tmp_path, "expected.xlsx", {"MissingSheet": {"A1": "x"}})

        result = compare_workbooks(actual, expected)

        assert result.matches is False
        assert any("missing" in d.lower() for d in result.diffs)

    def test_extra_sheet_in_actual(self, tmp_path):
        """異常系: actual に余分なシートが存在する場合は差分を検出する"""
        actual_data = {
            "Sheet1": {"A1": "value"},
            "ExtraSheet": {"A1": "extra"},
        }
        expected_data = {"Sheet1": {"A1": "value"}}
        actual = _make_workbook(tmp_path, "actual.xlsx", actual_data)
        expected = _make_workbook(tmp_path, "expected.xlsx", expected_data)

        result = compare_workbooks(actual, expected, sheets=None)

        assert result.matches is False
        assert any("ExtraSheet" in d for d in result.diffs)

    def test_excluded_cells_are_skipped(self, tmp_path):
        """正常系: excluded_cells に指定したセルは比較をスキップする"""
        actual = _make_workbook(tmp_path, "actual.xlsx", {"Sheet1": {"A1": "different"}})
        expected = _make_workbook(tmp_path, "expected.xlsx", {"Sheet1": {"A1": "original"}})
        excluded = [{"sheet": "Sheet1", "cells": ["A1"]}]

        result = compare_workbooks(actual, expected, excluded_cells=excluded)

        assert result.matches is True

    def test_specific_sheets_only(self, tmp_path):
        """正常系: sheets 指定により対象外シートの差分を無視する"""
        actual_data = {
            "Sheet1": {"A1": "same"},
            "Sheet2": {"A1": "different"},
        }
        expected_data = {
            "Sheet1": {"A1": "same"},
            "Sheet2": {"A1": "original"},
        }
        actual = _make_workbook(tmp_path, "actual.xlsx", actual_data)
        expected = _make_workbook(tmp_path, "expected.xlsx", expected_data)

        result = compare_workbooks(actual, expected, sheets=["Sheet1"])

        assert result.matches is True

    def test_invalid_file_path(self, tmp_path):
        """異常系: 存在しないファイルパスを指定した場合は matches=False を返す"""
        result = compare_workbooks(
            str(tmp_path / "nonexistent_actual.xlsx"),
            str(tmp_path / "nonexistent_expected.xlsx"),
        )

        assert result.matches is False
        assert len(result.diffs) > 0


# ---------------------------------------------------------------------------
# print_diff_report
# ---------------------------------------------------------------------------

class TestPrintDiffReport:
    """print_diff_report のテスト"""

    def test_no_differences(self, capsys):
        """正常系: 差分なしの場合は一致メッセージを出力する"""
        print_diff_report(DiffResult(matches=True, diffs=[]))
        captured = capsys.readouterr()
        assert "No differences" in captured.out

    def test_with_differences(self, capsys):
        """正常系: 差分ありの場合は差分件数と内容を出力する"""
        diffs = ["[Sheet1!A1] actual='x' != expected='y'"]
        print_diff_report(DiffResult(matches=False, diffs=diffs))
        captured = capsys.readouterr()
        assert "1" in captured.out
        assert "Sheet1" in captured.out
