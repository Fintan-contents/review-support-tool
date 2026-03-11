"""
xlsx_assertions モジュールのユニットテスト

get_result_sheet, get_detail_rows, has_error_sheet, count_review_record_issues,
get_header_info, get_category_headers, get_category_counts, get_review_record_issues
の正常系・異常系を検証する。
"""
import openpyxl
import pytest

from helpers.xlsx_assertions import (
    count_review_record_issues,
    get_category_counts,
    get_category_headers,
    get_detail_rows,
    get_header_info,
    get_result_sheet,
    get_review_record_issues,
    has_error_sheet,
)


# ---------------------------------------------------------------------------
# ユーティリティ
# ---------------------------------------------------------------------------

def _create_review_result_workbook(review_times: int = 1) -> openpyxl.Workbook:
    """レビュー結果シートを含むワークブックを作成するユーティリティ"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_name = f"レビュー結果{review_times}回目"
    wb.create_sheet(title=sheet_name)
    return wb


# ---------------------------------------------------------------------------
# get_result_sheet
# ---------------------------------------------------------------------------

class TestGetResultSheet:
    """get_result_sheet のテスト"""

    def test_sheet_exists(self):
        """正常系: 指定回数のレビュー結果シートを取得できる"""
        wb = _create_review_result_workbook(review_times=1)
        ws = get_result_sheet(wb, review_times=1)
        assert ws.title == "レビュー結果1回目"

    def test_sheet_exists_second_review(self):
        """正常系: 2回目のレビュー結果シートを取得できる"""
        wb = _create_review_result_workbook(review_times=2)
        ws = get_result_sheet(wb, review_times=2)
        assert ws.title == "レビュー結果2回目"

    def test_sheet_not_found(self):
        """異常系: 指定回数のシートが存在しない場合 AssertionError を送出する"""
        wb = _create_review_result_workbook(review_times=1)
        with pytest.raises(AssertionError):
            get_result_sheet(wb, review_times=99)


# ---------------------------------------------------------------------------
# get_detail_rows
# ---------------------------------------------------------------------------

class TestGetDetailRows:
    """get_detail_rows のテスト"""

    def _make_ws_with_rows(self) -> openpyxl.worksheet.worksheet.Worksheet:
        """指摘詳細行を持つワークシートを作成"""
        wb = openpyxl.Workbook()
        ws = wb.active
        # 行15から指摘詳細を配置
        ws.cell(15, 1).value = "設計書.xlsx"   # シート名
        ws.cell(15, 2).value = "B5"            # 位置（空でなければ有効行）
        ws.cell(15, 3).value = "山田太郎"       # レビュア
        ws.cell(15, 4).value = "a"             # カテゴリ
        ws.cell(15, 5).value = "内容が不明確"   # コメント
        ws.cell(15, 11).value = "未"           # 対応状況
        return ws

    def test_with_data_rows(self):
        """正常系: 有効な指摘行を辞書リストとして取得できる"""
        ws = self._make_ws_with_rows()
        rows = get_detail_rows(ws, start_row=15)

        assert len(rows) == 1
        assert rows[0]["position"] == "B5"
        assert rows[0]["reviewer"] == "山田太郎"
        assert rows[0]["category"] == "a"
        assert rows[0]["fix_status"] == "未"

    def test_empty_position_skipped(self):
        """正常系: 位置（セル位置）が空の行はスキップされる"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(15, 2).value = None  # 位置が空
        ws.cell(16, 2).value = "C3"  # 有効行

        rows = get_detail_rows(ws, start_row=15)

        assert len(rows) == 1
        assert rows[0]["position"] == "C3"

    def test_empty_sheet(self):
        """正常系: データなしシートは空リストを返す"""
        wb = openpyxl.Workbook()
        ws = wb.active
        rows = get_detail_rows(ws, start_row=15)
        assert rows == []


# ---------------------------------------------------------------------------
# has_error_sheet
# ---------------------------------------------------------------------------

class TestHasErrorSheet:
    """has_error_sheet のテスト"""

    def test_error_sheet_exists(self):
        """正常系: エラーシートが存在する場合 True を返す"""
        wb = openpyxl.Workbook()
        wb.create_sheet("エラーシート")
        assert has_error_sheet(wb) is True

    def test_error_sheet_not_exists(self):
        """正常系: エラーシートが存在しない場合 False を返す"""
        wb = openpyxl.Workbook()
        assert has_error_sheet(wb) is False


# ---------------------------------------------------------------------------
# count_review_record_issues
# ---------------------------------------------------------------------------

class TestCountReviewRecordIssues:
    """count_review_record_issues のテスト"""

    def test_count_valid_rows(self):
        """正常系: B列に値がある有効行数を正しくカウントする"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("レビュー指摘一覧")
        # ヘッダー行（スキップ: 1-4行目）
        # 有効行
        ws.cell(5, 2).value = 1   # B5: レビュー回数あり
        ws.cell(6, 2).value = 1   # B6: レビュー回数あり
        ws.cell(7, 2).value = None # B7: 空行（カウント対象外）

        count = count_review_record_issues(wb)
        assert count == 2

    def test_end_marker_stops_count(self):
        """正常系: A列に "END" マーカーがある行でカウントを停止する"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("レビュー指摘一覧")
        ws.cell(5, 2).value = 1     # 有効行
        ws.cell(6, 1).value = "END" # ENDマーカー
        ws.cell(6, 2).value = 1     # ENDより後（カウント対象外）
        ws.cell(7, 2).value = 1

        count = count_review_record_issues(wb)
        assert count == 1

    def test_sheet_not_exists(self):
        """正常系: 指定シートが存在しない場合 0 を返す"""
        wb = openpyxl.Workbook()
        count = count_review_record_issues(wb, sheet_name="存在しないシート")
        assert count == 0


# ---------------------------------------------------------------------------
# get_header_info
# ---------------------------------------------------------------------------

class TestGetHeaderInfo:
    """get_header_info のテスト"""

    def test_header_values(self):
        """正常系: レビュー結果シートのヘッダー情報を正しく取得できる"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(4, 1).value = "設計書.xlsx"  # A4: 対象ファイル
        ws.cell(4, 3).value = 1              # C4: レビュー回数
        ws.cell(4, 5).value = "2024-01-15"  # E4: 実施日

        info = get_header_info(ws)

        assert info["target_file"] == "設計書.xlsx"
        assert info["review_count"] == 1
        assert info["review_date"] == "2024-01-15"

    def test_empty_header(self):
        """正常系: ヘッダーが空の場合は None を返す"""
        wb = openpyxl.Workbook()
        ws = wb.active
        info = get_header_info(ws)
        assert info["target_file"] is None
        assert info["review_count"] is None


# ---------------------------------------------------------------------------
# get_category_headers
# ---------------------------------------------------------------------------

class TestGetCategoryHeaders:
    """get_category_headers のテスト"""

    def test_with_categories(self):
        """正常系: カテゴリ略称とカテゴリ名を正しく取得できる"""
        wb = openpyxl.Workbook()
        ws = wb.active
        # 7行目: カテゴリ略称
        ws.cell(7, 2).value = "a"
        ws.cell(7, 3).value = "b"
        # 8行目: カテゴリ名
        ws.cell(8, 2).value = "01_要件漏れ"
        ws.cell(8, 3).value = "02_仕様違反"

        result = get_category_headers(ws)

        assert result["abbreviations"] == ["a", "b"]
        assert result["names"] == ["01_要件漏れ", "02_仕様違反"]

    def test_empty_sheet(self):
        """正常系: カテゴリが設定されていない場合は空リストを返す"""
        wb = openpyxl.Workbook()
        ws = wb.active
        result = get_category_headers(ws)
        assert result["abbreviations"] == []
        assert result["names"] == []


# ---------------------------------------------------------------------------
# get_category_counts
# ---------------------------------------------------------------------------

class TestGetCategoryCounts:
    """get_category_counts のテスト"""

    def test_with_counts(self):
        """正常系: 済・未の件数を正しく取得できる"""
        wb = openpyxl.Workbook()
        ws = wb.active
        # カテゴリ略称（7行目）
        ws.cell(7, 2).value = "a"
        ws.cell(7, 3).value = "b"
        # 9行目: 済
        ws.cell(9, 2).value = 3
        ws.cell(9, 3).value = 1
        ws.cell(9, 11).value = 4   # 合計
        # 10行目: 未
        ws.cell(10, 2).value = 2
        ws.cell(10, 11).value = 2  # 合計

        result = get_category_counts(ws)

        assert result["fixed"] == {"a": 3, "b": 1}
        assert result["unfixed"] == {"a": 2}
        assert result["fixed_total"] == 4
        assert result["unfixed_total"] == 2


# ---------------------------------------------------------------------------
# get_review_record_issues
# ---------------------------------------------------------------------------

class TestGetReviewRecordIssues:
    """get_review_record_issues のテスト"""

    def test_with_issues(self):
        """正常系: レビュー記録票の指摘一覧を辞書リストとして取得できる"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("レビュー指摘一覧")
        # 5行目: 有効行
        ws.cell(5, 2).value = 1          # B: レビュー回数
        ws.cell(5, 3).value = "A1"       # C: 位置
        ws.cell(5, 4).value = "a"        # D: カテゴリ
        ws.cell(5, 5).value = "山田"     # E: レビュア
        ws.cell(5, 7).value = "未"       # G: 対応状況
        ws.cell(5, 8).value = "指摘内容" # H: 指摘コメント

        issues = get_review_record_issues(wb)

        assert len(issues) == 1
        assert issues[0]["review_count"] == 1
        assert issues[0]["position"] == "A1"
        assert issues[0]["category"] == "a"
        assert issues[0]["fix_status"] == "未"

    def test_end_marker_stops_processing(self):
        """正常系: A列の END マーカーでデータ取得を停止する"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        ws = wb.create_sheet("レビュー指摘一覧")
        ws.cell(5, 2).value = 1
        ws.cell(6, 1).value = "END"
        ws.cell(6, 2).value = 1  # END 以降は無視

        issues = get_review_record_issues(wb)
        assert len(issues) == 1

    def test_sheet_not_exists(self):
        """正常系: 指定シートが存在しない場合は空リストを返す"""
        wb = openpyxl.Workbook()
        issues = get_review_record_issues(wb, sheet_name="存在しないシート")
        assert issues == []
