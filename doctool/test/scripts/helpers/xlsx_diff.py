"""Excel ワークブック差分比較モジュール

期待結果xlsx（Gold Master）と実際の出力xlsxを比較し、差分を検出する。
"""
import openpyxl
from dataclasses import dataclass
from datetime import datetime
from typing import Optional


@dataclass
class DiffResult:
    """差分比較結果"""
    matches: bool
    diffs: list[str]


def compare_cells(actual, expected) -> bool:
    """型を考慮したセル値の比較
    
    Args:
        actual: 実際のセル値
        expected: 期待されるセル値
        
    Returns:
        bool: 値が一致する場合True
    """
    # 両方Noneの場合は一致
    if actual is None and expected is None:
        return True
    
    # どちらか一方のみNoneの場合は不一致
    if actual is None or expected is None:
        return False
    
    # datetime型の比較（1秒未満の誤差を許容）
    if isinstance(actual, datetime) and isinstance(expected, datetime):
        return abs((actual - expected).total_seconds()) < 1
    
    # 数値型の比較（浮動小数点誤差を考慮）
    if isinstance(actual, (int, float)) and isinstance(expected, (int, float)):
        return abs(float(actual) - float(expected)) < 1e-9
    
    # その他は文字列として比較
    return str(actual).strip() == str(expected).strip()


def _build_excluded_set(excluded_cells: list[dict]) -> dict[str, set[str]]:
    """excluded_cells 設定をシート名 → セル座標のセットに変換

    Args:
        excluded_cells: config.yaml の excluded_cells リスト
            例: [{"sheet": "レビュー結果1回目", "cells": ["E4", "F4"]}]

    Returns:
        dict: {シート名: {座標, ...}}
    """
    result = {}
    for entry in excluded_cells or []:
        sheet = entry.get("sheet", "")
        cells = set(entry.get("cells", []))
        if sheet:
            result.setdefault(sheet, set()).update(cells)
    return result


def _normalize_fill(cell) -> str:
    """セルの塗りつぶし色を正規化文字列で返す。塗りつぶしなし/透明の場合は空文字。

    テーマカラー等、openpyxl が正常に読み取れない色は空文字として扱い比較をスキップする。
    """
    fill = cell.fill
    if fill is None:
        return ""
    pt = getattr(fill, "patternType", None)
    if not pt or pt == "none":
        return ""
    if pt == "solid":
        try:
            rgb = getattr(fill.fgColor, "rgb", None)
        except Exception:
            return ""  # テーマカラー等、読み取り不可の場合はスキップ
        if not rgb or not isinstance(rgb, str):
            return ""
        # 8桁 RRGGBBAA 形式以外（テーマカラーのエラー文字列等）はスキップ
        import re as _re
        if not _re.fullmatch(r"[0-9A-Fa-f]{8}", rgb) or rgb == "00000000":
            return ""
        return f"solid:{rgb}"
    return pt


def _normalize_border_side(side) -> str:
    """罫線の一辺（left/right/top/bottom）を正規化文字列で返す。罫線なしは空文字。

    色は比較対象外とし、スタイル（thin/medium/thick など）のみを返す。
    テーマカラーやインデックスカラーは rgb 取得時に例外が発生するケースがあるため除外。
    """
    if side is None:
        return ""
    return getattr(side, "border_style", None) or ""


def _normalize_border(cell) -> str:
    """セルの4辺罫線を正規化文字列で返す。罫線が1辺もない場合は空文字。"""
    b = cell.border
    if b is None:
        return ""
    sides = {
        "l": _normalize_border_side(getattr(b, "left", None)),
        "r": _normalize_border_side(getattr(b, "right", None)),
        "t": _normalize_border_side(getattr(b, "top", None)),
        "b": _normalize_border_side(getattr(b, "bottom", None)),
    }
    parts = [f"{k}:{v}" for k, v in sides.items() if v]
    return ",".join(parts)


def _get_col_widths(ws) -> dict[str, float]:
    """明示的に設定された列幅を {列文字: 幅} で返す。幅未設定列は含まない。"""
    widths = {}
    for col_letter, col_dim in ws.column_dimensions.items():
        if col_dim.width is not None:
            widths[col_letter] = round(col_dim.width, 1)
    return widths


def compare_workbooks(
    actual_path: str,
    expected_path: str,
    sheets: Optional[list[str]] = None,
    excluded_cells: Optional[list[dict]] = None,
    ignore_format: bool = True,
    compare_fill: bool = False,
    compare_col_widths: bool = False,
    compare_print_area: bool = False,
    compare_borders: bool = False,
    max_diffs: int = 10
) -> DiffResult:
    """2つのxlsxファイルをセル値ベースで比較する

    Args:
        actual_path: 実際の出力xlsxファイルパス
        expected_path: 期待結果xlsxファイルパス
        sheets: 比較対象シート名リスト（Noneの場合は期待結果の全シートを比較）
        excluded_cells: 比較をスキップするセルの定義
            例: [{"sheet": "レビュー結果1回目", "cells": ["E4"]}]
            毎回変わる実施日時などのセルに使用する
        ignore_format: フォーマット（色・罫線等）を無視（現在は常にTrue）
        compare_fill: Trueの場合、セルの背景色（塗りつぶし）も比較する
        compare_col_widths: Trueの場合、明示設定された列幅を比較する
        compare_print_area: Trueの場合、印刷範囲を比較する
        compare_borders: Trueの場合、セルの4辺罫線（スタイル・色）を比較する
        max_diffs: 報告する差分の最大数（デバッグ容易性のため制限）

    Returns:
        DiffResult: 比較結果（一致/不一致 + 差分詳細リスト）
    """
    try:
        actual = openpyxl.load_workbook(actual_path, data_only=True)
        expected = openpyxl.load_workbook(expected_path, data_only=True)
    except Exception as e:
        return DiffResult(matches=False, diffs=[f"Failed to load workbooks: {e}"])

    diffs = []
    excluded = _build_excluded_set(excluded_cells)

    # 比較対象シートの決定（Noneの場合は期待結果の全シート）
    target_sheets = sheets if sheets is not None else expected.sheetnames

    for sheet_name in target_sheets:
        # シート存在チェック
        if sheet_name not in actual.sheetnames:
            diffs.append(f"Sheet '{sheet_name}' missing in actual workbook")
            if len(diffs) >= max_diffs:
                break
            continue

        if sheet_name not in expected.sheetnames:
            diffs.append(f"Sheet '{sheet_name}' missing in expected workbook")
            if len(diffs) >= max_diffs:
                break
            continue

        aws = actual[sheet_name]
        ews = expected[sheet_name]
        excluded_coords = excluded.get(sheet_name, set())

        # 行数・列数の比較
        if aws.max_row != ews.max_row:
            diffs.append(
                f"[{sheet_name}] Row count mismatch: "
                f"actual={aws.max_row}, expected={ews.max_row}"
            )

        if aws.max_column != ews.max_column:
            diffs.append(
                f"[{sheet_name}] Column count mismatch: "
                f"actual={aws.max_column}, expected={ews.max_column}"
            )

        # セル値・コメントの完全一致比較
        max_row = max(aws.max_row or 0, ews.max_row or 0)
        max_col = max(aws.max_column or 0, ews.max_column or 0)

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell_ref = aws.cell(row, col).coordinate

                # excluded_cells に指定されたセルはスキップ
                if cell_ref in excluded_coords:
                    continue

                av = aws.cell(row, col).value
                ev = ews.cell(row, col).value

                if not compare_cells(av, ev):
                    av_str = str(av)[:50] + "..." if av and len(str(av)) > 50 else str(av)
                    ev_str = str(ev)[:50] + "..." if ev and len(str(ev)) > 50 else str(ev)
                    diffs.append(
                        f"[{sheet_name}!{cell_ref}] actual='{av_str}' != expected='{ev_str}'"
                    )

                    if len(diffs) >= max_diffs:
                        diffs.append(f"... (差分数が{max_diffs}件を超えたため省略)")
                        return DiffResult(matches=False, diffs=diffs)

                # 背景色（塗りつぶし）の比較
                if compare_fill:
                    af = _normalize_fill(aws.cell(row, col))
                    ef = _normalize_fill(ews.cell(row, col))
                    if af != ef:
                        diffs.append(
                            f"[{sheet_name}!{cell_ref}] fill:"
                            f" actual='{af}' != expected='{ef}'"
                        )
                        if len(diffs) >= max_diffs:
                            diffs.append(f"... (差分数が{max_diffs}件を超えたため省略)")
                            return DiffResult(matches=False, diffs=diffs)

                # コメント（メモ）の比較
                ac = aws.cell(row, col).comment
                ec = ews.cell(row, col).comment
                ac_text = ac.text.strip() if ac else None
                ec_text = ec.text.strip() if ec else None
                if ac_text != ec_text:
                    ac_repr = repr(ac_text[:50] + "...") if ac_text and len(ac_text) > 50 else repr(ac_text)
                    ec_repr = repr(ec_text[:50] + "...") if ec_text and len(ec_text) > 50 else repr(ec_text)
                    diffs.append(
                        f"[{sheet_name}!{cell_ref}] comment: actual={ac_repr} != expected={ec_repr}"
                    )

                    if len(diffs) >= max_diffs:
                        diffs.append(f"... (差分数が{max_diffs}件を超えたため省略)")
                        return DiffResult(matches=False, diffs=diffs)

                # 罫線の比較
                if compare_borders:
                    ab = _normalize_border(aws.cell(row, col))
                    eb = _normalize_border(ews.cell(row, col))
                    if ab != eb:
                        diffs.append(
                            f"[{sheet_name}!{cell_ref}] border:"
                            f" actual='{ab}' != expected='{eb}'"
                        )
                        if len(diffs) >= max_diffs:
                            diffs.append(f"... (差分数が{max_diffs}件を超えたため省略)")
                            return DiffResult(matches=False, diffs=diffs)

        if len(diffs) >= max_diffs:
            diffs.append(f"... (差分数が{max_diffs}件を超えたため省略)")
            break

        # 列幅の比較（明示設定列のみ）
        if compare_col_widths and len(diffs) < max_diffs:
            aw_widths = _get_col_widths(aws)
            ew_widths = _get_col_widths(ews)
            all_cols = sorted(set(aw_widths) | set(ew_widths))
            for col_letter in all_cols:
                aw_w = aw_widths.get(col_letter)
                ew_w = ew_widths.get(col_letter)
                # expected側が設定済みの列のみ比較（expected未設定はデフォルト幅なので無視）
                if ew_w is not None:
                    actual_w = aw_w if aw_w is not None else 0.0
                    if abs(actual_w - ew_w) > 0.2:
                        diffs.append(
                            f"[{sheet_name}] Column {col_letter} width:"
                            f" actual={actual_w} != expected={ew_w}"
                        )
                        if len(diffs) >= max_diffs:
                            diffs.append(f"... (差分数が{max_diffs}件を超えたため省略)")
                            break
            else:
                if ew_widths:
                    print(f"  ✓ [{sheet_name}] column widths OK")

        # 印刷範囲の比較
        if compare_print_area and len(diffs) < max_diffs:
            actual_pa = str(aws.print_area) if aws.print_area else ""
            expected_pa = str(ews.print_area) if ews.print_area else ""
            if actual_pa != expected_pa:
                diffs.append(
                    f"[{sheet_name}] Print area: actual='{actual_pa}' != expected='{expected_pa}'"
                )
            elif expected_pa:
                print(f"  ✓ [{sheet_name}] print_area = '{expected_pa}'")

    # sheets=None（全シート比較モード）のとき、
    # actual にあって expected にない余分なシートを検出する
    # （削除されるべきシートが残存している場合を捕捉）
    if sheets is None and len(diffs) < max_diffs:
        extra = [s for s in actual.sheetnames if s not in expected.sheetnames]
        for sheet_name in extra:
            diffs.append(f"Sheet '{sheet_name}' exists in actual but not in expected")
            if len(diffs) >= max_diffs:
                diffs.append(f"... (差分数が{max_diffs}件を超えたため省略)")
                break

    return DiffResult(matches=len(diffs) == 0, diffs=diffs)


def print_diff_report(result: DiffResult) -> None:
    """差分レポートをコンソールに出力
    
    Args:
        result: 差分比較結果
    """
    if result.matches:
        print("✓ No differences found")
    else:
        print(f"✗ Found {len(result.diffs)} difference(s):")
        for i, diff in enumerate(result.diffs, 1):
            print(f"  {i}. {diff}")
