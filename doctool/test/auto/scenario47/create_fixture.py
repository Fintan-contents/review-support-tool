"""
S47 テストフィクスチャ生成スクリプト（複数設計書同時処理の集計検証）

2つの設計書（A・B）を同時に開いて抽出し、
レビュー記録一覧に2行が正しく追加されることを検証する。

【フィクスチャの設計】
  設計書A（システム機能設計書_サンプルA_S47.xlsx）:
    - カテゴリ a: 1件（未済）
    - カテゴリ c: 1件（未済）
    - → レビュー記録一覧 行1: a未=1, b未=0, c未=1

  設計書B（システム機能設計書_サンプルB_S47.xlsx）:
    - カテゴリ b: 1件（未済）
    - → レビュー記録一覧 行2: a未=0, b未=1, c未=0

【Gold Master 作成手順】
このスクリプトは入力フィクスチャのみ生成する。
Gold Master は以下の手順で作成:
  1. build.bat で xlsm をビルド
  2. Windows 環境で S47 の全ファイルを開いて抽出を実行
  3. 出力ファイルを _expected.xlsx としてコピーする

【実行方法】
    cd /path/to/review-support-tool/doctool/test
    python3 auto/scenario47/create_fixture.py
"""

import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.comments import Comment

SCRIPT_DIR = Path(__file__).parent
TEST_DIR = SCRIPT_DIR.parent.parent  # doctool/test
S01_DIR = TEST_DIR / "auto" / "scenario01"
S12_DIR = TEST_DIR / "auto" / "scenario12"
S29_DIR = TEST_DIR / "auto" / "scenario29"
S47_DIR = SCRIPT_DIR


def remove_existing_comments(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment:
                cell.comment = None


def add_comment(ws, cell_ref: str, text: str, author: str = "山田　太郎") -> None:
    ws[cell_ref].comment = Comment(text, author)


def create_design_doc_a(src_path: Path, dst_path: Path) -> None:
    """設計書A: カテゴリ a(未済) 1件 + c(未済) 1件"""
    wb = load_workbook(src_path)

    sheet_names = wb.sheetnames
    excluded = {"表紙", "変更履歴", "目次", "データ"}
    target_sheets = [s for s in sheet_names if s not in excluded]
    if not target_sheets:
        raise RuntimeError("コメント挿入可能なシートが存在しない")

    # 全シートの既存コメントを削除
    for sn in sheet_names:
        remove_existing_comments(wb[sn])

    ws0 = wb[target_sheets[0]]

    # 日時コメント
    add_comment(
        ws0, "A4",
        "山田　太郎:*\n実施日:2023/08/01\n開始:09:00\n終了:11:00\nレビュー時間:120",
    )
    # a カテゴリ: 1件（未済）
    add_comment(ws0, "B5", "山田　太郎:a\n設計書Aのaカテゴリ指摘。")
    # c カテゴリ: 1件（未済）
    add_comment(ws0, "B6", "山田　太郎:c\n設計書Aのcカテゴリ指摘。")

    wb.save(dst_path)
    print(f"Created: {dst_path.name} (a:1未, c:1未)")


def create_design_doc_b(src_path: Path, dst_path: Path) -> None:
    """設計書B: カテゴリ b(未済) 1件"""
    wb = load_workbook(src_path)

    sheet_names = wb.sheetnames
    excluded = {"表紙", "変更履歴", "目次", "データ"}
    target_sheets = [s for s in sheet_names if s not in excluded]
    if not target_sheets:
        raise RuntimeError("コメント挿入可能なシートが存在しない")

    for sn in sheet_names:
        remove_existing_comments(wb[sn])

    ws0 = wb[target_sheets[0]]

    # 日時コメント
    add_comment(
        ws0, "A4",
        "山田　太郎:*\n実施日:2023/08/02\n開始:10:00\n終了:12:00\nレビュー時間:120",
    )
    # b カテゴリ: 1件（未済）
    add_comment(ws0, "B5", "山田　太郎:b\n設計書Bのbカテゴリ指摘。")

    wb.save(dst_path)
    print(f"Created: {dst_path.name} (b:1未)")


def main():
    # S01 の設計書をベースとして使用（シート構成が判明している）
    src_design = S01_DIR / "システム機能設計書_サンプル_S01.xlsx"
    if not src_design.exists():
        # S01 が見つからない場合は S29 を使用
        src_design = S29_DIR / "システム機能設計書_サンプル_S29.xlsx"

    # ---- 1. 設計書A: a(未) 1件 + c(未) 1件 ----
    dst_design_a = S47_DIR / "システム機能設計書_サンプルA_S47.xlsx"
    create_design_doc_a(src_design, dst_design_a)

    # ---- 2. 設計書B: b(未) 1件 ----
    dst_design_b = S47_DIR / "システム機能設計書_サンプルB_S47.xlsx"
    create_design_doc_b(src_design, dst_design_b)

    # ---- 3. レビュー記録票A テンプレートコピー ----
    src_record = S01_DIR / "システム機能設計書_サンプル_レビュー記録票_S01.xlsx"
    if not src_record.exists():
        src_record = S29_DIR / "システム機能設計書_サンプル_S29_レビュー記録票.xlsx"
    dst_record_a = S47_DIR / "システム機能設計書_サンプルA_S47_レビュー記録票.xlsx"
    shutil.copy2(src_record, dst_record_a)
    print(f"Copied: {dst_record_a.name}")

    # ---- 4. レビュー記録票B テンプレートコピー ----
    dst_record_b = S47_DIR / "システム機能設計書_サンプルB_S47_レビュー記録票.xlsx"
    shutil.copy2(src_record, dst_record_b)
    print(f"Copied: {dst_record_b.name}")

    # ---- 5. レビュー記録サマリ テンプレートコピー ----
    src_summary = S29_DIR / "レビュー記録サマリ_S29.xlsx"
    dst_summary = S47_DIR / "レビュー記録サマリ_S47.xlsx"
    shutil.copy2(src_summary, dst_summary)
    print(f"Copied: {dst_summary.name}")

    print("\nDone. Input fixture files created:")
    for f in sorted(S47_DIR.glob("*.xlsx")):
        print(f"  {f.name}")
    print("\nExpected Gold Master:")
    print("  レビュー記録サマリ: 2行追加")
    print("    行1 (設計書A): a未=1, b未=0, c未=1")
    print("    行2 (設計書B): a未=0, b未=1, c未=0")
    print("\nNOTE: Gold Master (_expected.xlsx) must be created on Windows.")


if __name__ == "__main__":
    main()
