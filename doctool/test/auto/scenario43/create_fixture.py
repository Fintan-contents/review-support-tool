"""
S43 テストフィクスチャ生成スクリプト（バックアップ機能検証）

CbBkup=True を設定した状態で抽出を実行し、
レビュー記録一覧のバックアップファイル（レビュー記録サマリ_S43_YYYYMMDD_HHMMSS.xlsx）が
作業ディレクトリに作成されることを検証する。

【目的】
OpenReviewListFile 内の FileCopy 呼び出しを検証する。
バックアップ成功を assert_file_glob_exists で確認するため、
入力フィクスチャは S42 と同じ「コメントなし・3カテゴリ」を使用する。

【Gold Master 作成手順】
このスクリプトは入力フィクスチャのみ生成する。
Gold Master は以下の手順で作成:
  1. build.bat で xlsm をビルド
  2. Windows 環境で S43 の設計書で抽出を実行（CbBkup=True）
  3. 出力ファイルを _expected.xlsx としてコピーする
  ※ バックアップファイルの存在は assert_file_glob_exists で自動検証される

【実行方法】
    cd /path/to/review-support-tool/doctool/test
    python3 auto/scenario43/create_fixture.py
"""

import shutil
from pathlib import Path
from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).parent
TEST_DIR = SCRIPT_DIR.parent.parent  # doctool/test
S29_DIR = TEST_DIR / "auto" / "scenario29"
S43_DIR = SCRIPT_DIR


def remove_existing_comments(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment:
                cell.comment = None


def main():
    # ---- 1. システム機能設計書: コメントをすべて削除した版を作成 ----
    src_design = S29_DIR / "システム機能設計書_サンプル_S29.xlsx"
    dst_design = S43_DIR / "システム機能設計書_サンプル_S43.xlsx"

    print(f"Loading: {src_design}")
    wb = load_workbook(src_design)

    for sheet_name in wb.sheetnames:
        remove_existing_comments(wb[sheet_name])

    print(f"Saving: {dst_design}")
    wb.save(dst_design)
    print("OK: All comments removed (total=0).")

    # ---- 2. レビュー記録サマリ テンプレートコピー ----
    src_summary = S29_DIR / "レビュー記録サマリ_S29.xlsx"
    dst_summary = S43_DIR / "レビュー記録サマリ_S43.xlsx"
    print(f"Copying summary template: {dst_summary}")
    shutil.copy2(src_summary, dst_summary)

    # ---- 3. レビュー記録票 テンプレートコピー ----
    src_record = S29_DIR / "システム機能設計書_サンプル_S29_レビュー記録票.xlsx"
    dst_record = S43_DIR / "システム機能設計書_サンプル_S43_レビュー記録票.xlsx"
    print(f"Copying review record template: {dst_record}")
    shutil.copy2(src_record, dst_record)

    print("\nDone. Input fixture files created:")
    for f in sorted(S43_DIR.glob("*.xlsx")):
        print(f"  {f.name}")
    print("\nNOTE:")
    print("  Gold Master (_expected.xlsx) must be created after VBA build on Windows.")
    print("  The backup file (レビュー記録サマリ_S43_YYYYMMDD_HHMMSS.xlsx) is validated")
    print("  automatically via assert_file_glob_exists in config.yaml.")


if __name__ == "__main__":
    main()
