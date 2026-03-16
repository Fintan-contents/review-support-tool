"""
S41 テストフィクスチャ生成スクリプト（条件付合格 検証）

CbConditional=True + CONDITIONAL_CATEGORY="j" + CONDITIONAL_COUNT=3 の設定で、
"j" カテゴリのコメントのみ2件（閾値3以下）を含む設計書を使用する。
reviewResult="条件付合格" が正しい列（DETAIL_COL_LIST_CATEGORY_A + 12 = 列41）に
書き込まれることを検証する。

【DetermineReviewResult の動作（S41 で期待される動作）】
  - CbConditional = True
  - j カテゴリの件数 = 2（> 0）
  - Range("CONDITIONAL_CATEGORY") = "j" → 一致
  - 件数(2) <= Range("CONDITIONAL_COUNT")(3) かつ errcount=0
  - → reviewResult = "条件付合格"

【Gold Master 作成手順】
このスクリプトは入力フィクスチャのみ生成する。
Gold Master は VBA 修正後に以下の手順で作成:
  1. VBA の DETAIL_COL_LIST_REVIEW_RESULT を動的計算式に修正してビルドする
  2. Windows 環境で S41 の設計書で抽出を実行する
  3. 出力ファイルを _expected.xlsx としてコピーする

【実行方法】
    cd /path/to/review-support-tool/doctool/test
    python3 auto/scenario41/create_fixture.py
"""

import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.comments import Comment

SCRIPT_DIR = Path(__file__).parent
TEST_DIR = SCRIPT_DIR.parent.parent  # doctool/test
S29_DIR = TEST_DIR / "auto" / "scenario29"
S41_DIR = SCRIPT_DIR

# j のみ2件（閾値3以下 → 条件付合格）。他カテゴリは0件
J_COMMENT_COUNT = 2


def remove_existing_comments(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment:
                cell.comment = None


def add_comment(ws, cell_ref: str, text: str, author: str = "山田　太郎") -> None:
    ws[cell_ref].comment = Comment(text, author)


def main():
    # ---- 1. システム機能設計書: j カテゴリのコメント2件のみ挿入 ----
    src_design = S29_DIR / "システム機能設計書_サンプル_S29.xlsx"
    dst_design = S41_DIR / "システム機能設計書_サンプル_S41.xlsx"

    print(f"Loading: {src_design}")
    wb = load_workbook(src_design)

    sheet_names = wb.sheetnames
    excluded = {"表紙", "変更履歴", "目次", "データ"}
    target_sheets = [s for s in sheet_names if s not in excluded]
    if not target_sheets:
        raise RuntimeError("コメント挿入可能なシートが存在しない")

    # 全シートの既存コメントを削除
    for sn in sheet_names:
        remove_existing_comments(wb[sn])

    ws0 = wb[target_sheets[0]]

    # 日時コメント（*カテゴリ）を挿入
    add_comment(
        ws0, "A4",
        "山田　太郎:*\n実施日:2023/08/01\n開始:09:00\n終了:11:00\nレビュー時間:120",
    )

    # j カテゴリのコメントのみ挿入（2件、閾値3以下 → 条件付合格）
    for idx in range(J_COMMENT_COUNT):
        row = 5 + idx
        cell_ref = f"B{row}"
        text = f"山田　太郎:j\njカテゴリの指摘内容サンプル{idx + 1}。"
        add_comment(ws0, cell_ref, text)

    print(f"Saving: {dst_design}")
    wb.save(dst_design)
    print(f"OK: 'j' category {J_COMMENT_COUNT} comments inserted (other categories=0).")
    print("CbConditional=True + CONDITIONAL_CATEGORY='j' + CONDITIONAL_COUNT=3")
    print("→ Expected reviewResult='条件付合格'")

    # ---- 2. レビュー記録サマリ テンプレートコピー ----
    src_summary = S29_DIR / "レビュー記録サマリ_S29.xlsx"
    dst_summary = S41_DIR / "レビュー記録サマリ_S41.xlsx"
    print(f"Copying summary template: {dst_summary}")
    shutil.copy2(src_summary, dst_summary)

    # ---- 3. レビュー記録票 テンプレートコピー ----
    src_record = S29_DIR / "システム機能設計書_サンプル_S29_レビュー記録票.xlsx"
    dst_record = S41_DIR / "システム機能設計書_サンプル_S41_レビュー記録票.xlsx"
    print(f"Copying review record template: {dst_record}")
    shutil.copy2(src_record, dst_record)

    print("\nDone. Input fixture files created:")
    for f in sorted(S41_DIR.glob("*.xlsx")):
        print(f"  {f.name}")
    print("\nNOTE: Gold Master (_expected.xlsx) must be created after VBA fix on Windows.")


if __name__ == "__main__":
    main()
