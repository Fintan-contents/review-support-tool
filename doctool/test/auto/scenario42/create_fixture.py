"""
S42 テストフィクスチャ生成スクリプト（3カテゴリ + 指摘ゼロ → 合格 検証）

コメントを含まない設計書（指摘合計=0）を使い、reviewResult="合格" が
正しい列（DETAIL_COL_LIST_CATEGORY_A + 3 = 列32）に書き込まれることを検証する。

【目的】
カテゴリ数が9未満の場合の REVIEW_RESULT 列位置正当性を検証する。
動的計算式 DETAIL_COL_LIST_CATEGORY_A + categoryMappings.Count は N=9 のとき
旧定数38と一致するため、N<9 のケースを個別に検証する必要がある。

3カテゴリの場合: REVIEW_RESULT = 29 + 3 = 列32（AF列）
固定定数 38 を使った場合: 列38（AL列）に書き込まれてしまう（バグ）

【Gold Master 作成手順】
このスクリプトは入力フィクスチャのみ生成する。
Gold Master は以下の手順で作成:
  1. build.bat で xlsm をビルド
  2. Windows 環境で S42 の設計書で抽出を実行（コメントなし → 合格が列32に入る）
  3. 出力ファイルを _expected.xlsx としてコピーする

【実行方法】
    cd /path/to/review-support-tool/doctool/test
    python3 auto/scenario42/create_fixture.py
"""

import shutil
from pathlib import Path
from openpyxl import load_workbook

SCRIPT_DIR = Path(__file__).parent
TEST_DIR = SCRIPT_DIR.parent.parent  # doctool/test
S29_DIR = TEST_DIR / "auto" / "scenario29"
S42_DIR = SCRIPT_DIR


def remove_existing_comments(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment:
                cell.comment = None


def main():
    # ---- 1. システム機能設計書: コメントをすべて削除した版を作成 ----
    src_design = S29_DIR / "システム機能設計書_サンプル_S29.xlsx"
    dst_design = S42_DIR / "システム機能設計書_サンプル_S42.xlsx"

    print(f"Loading: {src_design}")
    wb = load_workbook(src_design)

    # 全シートのコメントを削除（指摘ゼロ → reviewResult="合格" を実現）
    for sheet_name in wb.sheetnames:
        remove_existing_comments(wb[sheet_name])

    print(f"Saving: {dst_design}")
    wb.save(dst_design)
    print("OK: All comments removed (total=0 → reviewResult will be '合格').")

    # ---- 2. レビュー記録サマリ テンプレートコピー ----
    src_summary = S29_DIR / "レビュー記録サマリ_S29.xlsx"
    dst_summary = S42_DIR / "レビュー記録サマリ_S42.xlsx"
    print(f"Copying summary template: {dst_summary}")
    shutil.copy2(src_summary, dst_summary)

    # ---- 3. レビュー記録票 テンプレートコピー ----
    src_record = S29_DIR / "システム機能設計書_サンプル_S29_レビュー記録票.xlsx"
    dst_record = S42_DIR / "システム機能設計書_サンプル_S42_レビュー記録票.xlsx"
    print(f"Copying review record template: {dst_record}")
    shutil.copy2(src_record, dst_record)

    print("\nDone. Input fixture files created:")
    for f in sorted(S42_DIR.glob("*.xlsx")):
        print(f"  {f.name}")
    print("\nNOTE: Gold Master (_expected.xlsx) must be created after VBA build on Windows.")


if __name__ == "__main__":
    main()
