"""
S37 テストフィクスチャ生成スクリプト

12カテゴリ（a-l）を設定し、エイリアス "j"（10番目）のコメントを1件含む
入力フィクスチャを生成する。

【目的】
DETAIL_COL_LIST_REVIEW_RESULT 列位置バグの検証用シナリオ。
カテゴリ数が10以上のとき、固定定数 DETAIL_COL_LIST_REVIEW_RESULT = 38 により
エイリアス "j"（DETAIL_COL_LIST_CATEGORY_A + 9 = 列38）の件数が
reviewResult の値で上書きされるバグを検出する。

【Gold Master 作成手順】
このスクリプトは入力フィクスチャのみ生成する。
Gold Master (_expected.xlsx) は VBA 修正後に以下の手順で作成すること:
  1. VBA の DETAIL_COL_LIST_REVIEW_RESULT / DETAIL_COL_LIST_RE_REVIEW_WAY を
     動的計算式に修正してビルドする
  2. Windows 環境で doctool.xlsm を開き、S37 の設計書で抽出を実行する
  3. 出力された各ファイルを _expected.xlsx としてリネームしてコピーする:
     - システム機能設計書_サンプル_S37.xlsx → システム機能設計書_サンプル_S37_expected.xlsx
     - システム機能設計書_サンプル_S37_レビュー記録票.xlsx
         → システム機能設計書_サンプル_S37_レビュー記録票_expected.xlsx
     - レビュー記録サマリ_S37.xlsx → レビュー記録サマリ_S37_expected.xlsx

【実行方法（WSL2 / Windows 両環境で実行可能）】
    cd /path/to/review-support-tool/doctool/test
    source .venv/bin/activate
    python3 auto/scenario37/create_fixture.py
"""

import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.comments import Comment

SCRIPT_DIR = Path(__file__).parent
TEST_DIR = SCRIPT_DIR.parent.parent  # doctool/test
S29_DIR = TEST_DIR / "auto" / "scenario29"
S37_DIR = SCRIPT_DIR

CATEGORIES = ["a", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l"]
assert len(CATEGORIES) == 12, f"Expected 12 categories, got {len(CATEGORIES)}"


def add_comment(ws, cell_ref: str, text: str, author: str = "山田　太郎") -> None:
    ws[cell_ref].comment = Comment(text, author)


def remove_existing_comments(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment:
                cell.comment = None


def main():
    # ---- 1. システム機能設計書 フィクスチャ作成 ----
    src_design = S29_DIR / "システム機能設計書_サンプル_S29.xlsx"
    dst_design = S37_DIR / "システム機能設計書_サンプル_S37.xlsx"

    print(f"Loading: {src_design}")
    wb = load_workbook(src_design)

    sheet_names = wb.sheetnames
    print(f"Available sheets: {sheet_names}")

    excluded = {"表紙", "変更履歴", "目次", "データ"}
    target_sheets = [s for s in sheet_names if s not in excluded]
    print(f"Target sheets ({len(target_sheets)}): {target_sheets}")

    if len(target_sheets) < 1:
        raise RuntimeError("コメント挿入可能なシートが存在しない")

    # 各シートの既存コメントを削除
    for sn in target_sheets:
        remove_existing_comments(wb[sn])

    ws0 = wb[target_sheets[0]]

    # 日時コメント（*カテゴリ）を Sheet1 A4 に挿入
    add_comment(
        ws0, "A4",
        "山田　太郎:*\n実施日:2023/08/01\n開始:09:00\n終了:11:00\nレビュー時間:120",
    )

    # 12カテゴリのコメントを Sheet1 に挿入
    # 特に "j"（10番目、index=9）のコメントが列位置バグの検証対象
    for idx, alias in enumerate(CATEGORIES):
        row = 5 + idx
        cell_ref = f"B{row}"
        content = f"{alias}カテゴリの指摘内容サンプル。"
        text = f"山田　太郎:{alias}\n{content}"
        add_comment(ws0, cell_ref, text)

    print(f"Saving: {dst_design}")
    wb.save(dst_design)
    print("OK: 12-category fixture created (including alias 'j' at index 9).")

    # ---- 2. レビュー記録サマリ テンプレートコピー ----
    src_summary = S29_DIR / "レビュー記録サマリ_S29.xlsx"
    dst_summary = S37_DIR / "レビュー記録サマリ_S37.xlsx"
    print(f"Copying summary template: {dst_summary}")
    shutil.copy2(src_summary, dst_summary)

    # ---- 3. レビュー記録票 テンプレートコピー ----
    src_record = S29_DIR / "システム機能設計書_サンプル_S29_レビュー記録票.xlsx"
    dst_record = S37_DIR / "システム機能設計書_サンプル_S37_レビュー記録票.xlsx"
    print(f"Copying review record template: {dst_record}")
    shutil.copy2(src_record, dst_record)

    print("\nDone. Input fixture files created:")
    for f in sorted(S37_DIR.glob("*.xlsx")):
        print(f"  {f.name}")
    print("\nNOTE: Gold Master (_expected.xlsx) must be created after VBA fix on Windows.")
    print("See module docstring for instructions.")


if __name__ == "__main__":
    main()
