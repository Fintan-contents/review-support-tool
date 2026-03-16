"""
S40 テストフィクスチャ生成スクリプト（指摘ゼロ → 合格 検証）

コメントを含まない設計書（指摘合計=0）を使い、reviewResult="合格" が
正しい列（DETAIL_COL_LIST_CATEGORY_A + 12 = 列41）に書き込まれることを検証する。

【目的】
DETAIL_COL_LIST_REVIEW_RESULT 列位置バグの「合格」バリエーション検証。
固定定数 DETAIL_COL_LIST_REVIEW_RESULT = 38 の場合、
"合格" がエイリアス "j" の列（列38）に書き込まれてしまう。

【Gold Master 作成手順】
このスクリプトは入力フィクスチャのみ生成する。
Gold Master は VBA 修正後に以下の手順で作成:
  1. VBA の DETAIL_COL_LIST_REVIEW_RESULT / DETAIL_COL_LIST_RE_REVIEW_WAY を
     動的計算式に修正してビルドする
  2. Windows 環境で S40 の設計書で抽出を実行する（コメントなし → 合格が列41に入る）
  3. 出力ファイルを _expected.xlsx としてコピーする

【実行方法】
    cd /path/to/review-support-tool/doctool/test
    python3 auto/scenario40/create_fixture.py
"""

import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.comments import Comment

SCRIPT_DIR = Path(__file__).parent
TEST_DIR = SCRIPT_DIR.parent.parent  # doctool/test
S29_DIR = TEST_DIR / "auto" / "scenario29"
S40_DIR = SCRIPT_DIR


def remove_existing_comments(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment:
                cell.comment = None


def main():
    # ---- 1. システム機能設計書: コメントをすべて削除した版を作成 ----
    src_design = S29_DIR / "システム機能設計書_サンプル_S29.xlsx"
    dst_design = S40_DIR / "システム機能設計書_サンプル_S40.xlsx"

    print(f"Loading: {src_design}")
    wb = load_workbook(src_design)

    # 全シートのコメントを削除（指摘ゼロ → reviewResult="合格" を実現）
    for sheet_name in wb.sheetnames:
        remove_existing_comments(wb[sheet_name])

    print(f"Saving: {dst_design}")
    wb.save(dst_design)
    print("OK: All comments removed (total=0 → reviewResult will be '合格').")

    # ---- 2. レビュー記録サマリ テンプレートコピー ----
    src_summary = S29_DIR / "レビュー記録サマリ_S29.xlsx"
    dst_summary = S40_DIR / "レビュー記録サマリ_S40.xlsx"
    print(f"Copying summary template: {dst_summary}")
    shutil.copy2(src_summary, dst_summary)

    # ---- 3. レビュー記録票 テンプレートコピー ----
    src_record = S29_DIR / "システム機能設計書_サンプル_S29_レビュー記録票.xlsx"
    dst_record = S40_DIR / "システム機能設計書_サンプル_S40_レビュー記録票.xlsx"
    print(f"Copying review record template: {dst_record}")
    shutil.copy2(src_record, dst_record)

    print("\nDone. Input fixture files created:")
    for f in sorted(S40_DIR.glob("*.xlsx")):
        print(f"  {f.name}")
    print("\nNOTE: Gold Master (_expected.xlsx) must be created after VBA fix on Windows.")


if __name__ == "__main__":
    main()
