"""
S35 テストフィクスチャ生成スクリプト

S29 のフィクスチャをベースに、全52カテゴリ（a～z, aa～zz）のコメントを
3シートにまたがって挿入した S35 フィクスチャを生成する。

使用方法:
    cd /home/kuma/work/review-support-tool/doctool/test
    source .venv/bin/activate
    python3 auto/scenario35/create_fixture.py
"""

import shutil
from pathlib import Path
import openpyxl
from openpyxl.comments import Comment

SCRIPT_DIR = Path(__file__).parent
TEST_DIR = SCRIPT_DIR.parent.parent  # doctool/test
S29_DIR = TEST_DIR / "auto" / "scenario29"
S35_DIR = SCRIPT_DIR

# 全52カテゴリのエイリアス（a-z: 26, aa-zz: 26）
SINGLE_CATEGORIES = list("abcdefghijklmnopqrstuvwxyz")  # 26件
DOUBLE_CATEGORIES = [c * 2 for c in "abcdefghijklmnopqrstuvwxyz"]  # aa,bb,...,zz 26件
ALL_CATEGORIES = SINGLE_CATEGORIES + DOUBLE_CATEGORIES  # 52件


def create_comment_text(reviewer: str, alias: str, content: str, reply: str = None) -> str:
    """VBA が解析するコメントテキストを生成する。

    形式: "{reviewer}:{alias}\n{content}"
    返信あり: "{reviewer}:{alias}\n{content}\n---\n{reply}"
    """
    text = f"{reviewer}:{alias}\n{content}"
    if reply:
        text += f"\n---\n{reply}"
    return text


def add_comment(ws, cell_ref: str, text: str, author: str = "山田　太郎") -> None:
    """ワークシートのセルにコメントを追加する。"""
    comment = Comment(text, author)
    ws[cell_ref].comment = comment


def remove_existing_comments(ws) -> None:
    """ワークシートの既存コメントを全て削除する。"""
    for row in ws.iter_rows():
        for cell in row:
            if cell.comment:
                cell.comment = None


def build_cell_refs(start_row: int, count: int, col: str = "B") -> list:
    """連続するセル参照リストを生成する（例: B5, B6, B7...）。"""
    return [f"{col}{start_row + i}" for i in range(count)]


def main():
    # ---- 1. システム機能設計書 フィクスチャ作成 ----
    src_design = S29_DIR / "システム機能設計書_サンプル_S29.xlsx"
    dst_design = S35_DIR / "システム機能設計書_サンプル_S35.xlsx"

    print(f"Loading: {src_design}")
    wb = openpyxl.load_workbook(src_design)

    # 日時コメント挿入先シート（常に存在するシートに入れる）
    sheet_names = wb.sheetnames
    print(f"Available sheets: {sheet_names}")

    # コメント挿入対象シートを3枚選定（インデックス3,4,5あたり）
    target_sheets = []
    for name in sheet_names:
        if name not in ("表紙", "変更履歴", "目次", "データ"):
            target_sheets.append(name)
        if len(target_sheets) >= 3:
            break

    if len(target_sheets) < 3:
        raise RuntimeError(f"コメント挿入可能なシートが3枚未満: {target_sheets}")

    print(f"Target sheets: {target_sheets}")

    # 各シートの既存コメントを削除
    for sn in target_sheets:
        ws = wb[sn]
        remove_existing_comments(ws)

    # 52カテゴリを3シートに分散配置
    # Sheet1: 18カテゴリ（a-r）
    # Sheet2: 17カテゴリ（s-z, aa-hh）
    # Sheet3: 17カテゴリ（ii-zz）
    distribution = [
        (target_sheets[0], ALL_CATEGORIES[0:18]),
        (target_sheets[1], ALL_CATEGORIES[18:35]),
        (target_sheets[2], ALL_CATEGORIES[35:52]),
    ]

    # 日時コメントはSheet1に1件挿入（*カテゴリ）
    ws0 = wb[target_sheets[0]]
    add_comment(
        ws0, "A4",
        "山田　太郎:*\n実施日:2023/08/01\n開始:09:00\n終了:11:00\nレビュー時間:120",
        author="山田　太郎",
    )

    for sheet_name, categories in distribution:
        ws = wb[sheet_name]
        # カテゴリごとに異なる行に配置（5行目から2行おき）
        for idx, alias in enumerate(categories):
            row = 5 + idx * 2
            cell_ref = f"B{row}"
            content = f"{alias}カテゴリの指摘内容サンプル。設計書の記述を確認すること。"

            # 一部のコメント（先頭2件）に返信（済）を付ける
            if idx < 2:
                reply = f"対応しました。\n済"
                text = create_comment_text("山田　太郎", alias, content, reply=reply)
            else:
                text = create_comment_text("山田　太郎", alias, content)

            add_comment(ws, cell_ref, text, author="山田　太郎")

    print(f"Saving: {dst_design}")
    wb.save(dst_design)

    # ---- 2. カテゴリ総数の確認 ----
    wb_check = openpyxl.load_workbook(dst_design)
    total_comments = 0
    category_found = set()
    for sn in wb_check.sheetnames:
        ws = wb_check[sn]
        for row in ws.iter_rows():
            for cell in row:
                if cell.comment:
                    text = cell.comment.text
                    # カテゴリを抽出（コロン後の1-2文字）
                    colon_pos = text.find(":")
                    if colon_pos >= 0:
                        lf_pos = text.find("\n", colon_pos)
                        if lf_pos < 0:
                            lf_pos = len(text)
                        cat_len = min(lf_pos - colon_pos - 1, 2)
                        if cat_len >= 1:
                            cat = text[colon_pos + 1: colon_pos + 1 + cat_len]
                            if cat != "*":
                                category_found.add(cat)
                    total_comments += 1

    print(f"Total comments inserted: {total_comments}")
    print(f"Unique categories found: {len(category_found)}")
    missing = set(ALL_CATEGORIES) - category_found
    if missing:
        print(f"WARNING: Missing categories: {sorted(missing)}")
    else:
        print("OK: All 52 categories are present.")

    # ---- 3. レビュー記録サマリ テンプレートコピー ----
    src_summary = S29_DIR / "レビュー記録サマリ_S29.xlsx"
    dst_summary = S35_DIR / "レビュー記録サマリ_S35.xlsx"
    print(f"Copying summary template: {dst_summary}")
    shutil.copy2(src_summary, dst_summary)

    # ---- 4. レビュー記録票 テンプレートコピー ----
    src_record = S29_DIR / "システム機能設計書_サンプル_S29_レビュー記録票.xlsx"
    dst_record = S35_DIR / "システム機能設計書_サンプル_S35_レビュー記録票.xlsx"
    print(f"Copying review record template: {dst_record}")
    shutil.copy2(src_record, dst_record)

    print("\nDone. Fixture files created:")
    for f in S35_DIR.glob("*.xlsx"):
        print(f"  {f.name}")


if __name__ == "__main__":
    main()
